# create_calendar_header_2026.py
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

def create_calendar_header_2026():
    """
    Создает календарь в формате заголовков столбцов для Excel
    """
    
    # Создаем все дни 2026 года
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2026, 12, 31)
    
    all_dates = []
    current_date = start_date
    while current_date <= end_date:
        all_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Праздники 2026 (ст. 112 ТК РФ)
    holidays = [
        # Новогодние каникулы
        *[datetime(2026, 1, d) for d in range(1, 9)],
        datetime(2026, 1, 7),  # Рождество
        datetime(2026, 2, 23),  # 23 февраля
        datetime(2026, 3, 8),   # 8 марта
        datetime(2026, 5, 1),   # 1 мая
        datetime(2026, 5, 9),   # 9 мая
        datetime(2026, 6, 12),  # 12 июня
        datetime(2026, 11, 4),  # 4 ноября
    ]
    
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "График отпусков 2026"
    
    # Стили для форматирования
    # Рабочие дни
    workday_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # белый
    workday_font = Font(color="000000", bold=False)  # черный
    
    # Выходные
    weekend_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # светло-серый
    weekend_font = Font(color="000000", bold=False)  # черный
    
    # Праздники
    holiday_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # светло-красный
    holiday_font = Font(color="000000", bold=True)  # черный жирный
    
    # Предпраздничные
    preholiday_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # светло-желтый
    preholiday_font = Font(color="000000", italic=True)  # черный курсив
    
    # Заголовки столбцов
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # синий
    header_font = Font(color="FFFFFF", bold=True)  # белый жирный
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== СОЗДАЕМ ЗАГОЛОВОК КАЛЕНДАРЯ ==========
    
    # Первая строка: Месяцы и дни недели
    ws['A1'] = "№"
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].alignment = header_alignment
    ws['A1'].border = thin_border
    
    ws['B1'] = "ФИО сотрудника"
    ws['B1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].alignment = header_alignment
    ws['B1'].border = thin_border
    
    ws['C1'] = "Должность"
    ws['C1'].fill = header_fill
    ws['C1'].font = header_font
    ws['C1'].alignment = header_alignment
    ws['C1'].border = thin_border
    
    ws['D1'] = "Отдел"
    ws['D1'].fill = header_fill
    ws['D1'].font = header_font
    ws['D1'].alignment = header_alignment
    ws['D1'].border = thin_border
    
    # Заполняем календарь по столбцам, начиная с E1
    col_index = 5  # Начинаем с колонки E
    
    # Словарь для названий месяцев
    months_ru = {
        1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр", 5: "Май", 6: "Июн",
        7: "Июл", 8: "Авг", 9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек"
    }
    
    days_short = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    # Предпраздничные дни (сокращенный рабочий день)
    pre_holidays = [
        datetime(2026, 2, 20),  # перед 23 февраля
        datetime(2026, 3, 7),   # перед 8 марта
        datetime(2026, 5, 8),   # перед 9 мая
        datetime(2026, 6, 11),  # перед 12 июня
        datetime(2026, 11, 3),  # перед 4 ноября
        datetime(2026, 12, 31), # перед Новым годом
    ]
    
    current_month = 0
    month_cols = {}  # Для группировки колонок по месяцам
    
    for i, date in enumerate(all_dates, 1):
        col_letter = get_column_letter(col_index)
        
        # Определяем тип дня
        weekday = date.weekday()  # 0-пн, 6-вс
        is_weekend = weekday >= 5
        is_holiday = date in holidays
        is_preholiday = date in pre_holidays
        
        # Форматируем ячейку
        cell = ws[f"{col_letter}1"]
        
        if is_holiday:
            cell.value = f"{date.day}\n✶"
            cell.fill = holiday_fill
            cell.font = holiday_font
        elif is_preholiday:
            cell.value = f"{date.day}\n◐"
            cell.fill = preholiday_fill
            cell.font = preholiday_font
        elif is_weekend:
            cell.value = f"{date.day}\n{days_short[weekday]}"
            cell.fill = weekend_fill
            cell.font = weekend_font
        else:
            cell.value = f"{date.day}\n{days_short[weekday]}"
            cell.fill = workday_fill
            cell.font = workday_font
        
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Вторая строка: номер месяца
        cell2 = ws[f"{col_letter}2"]
        if date.month != current_month:
            cell2.value = months_ru[date.month]
            cell2.fill = header_fill
            cell2.font = header_font
            cell2.alignment = header_alignment
            cell2.border = thin_border
            current_month = date.month
            # Запоминаем начало месяца для объединения
            if date.month not in month_cols:
                month_cols[date.month] = col_index
        else:
            cell2.value = ""
        
        # Настраиваем ширину столбца
        ws.column_dimensions[col_letter].width = 4
        
        col_index += 1
    
    # Объединяем ячейки с названиями месяцев
    for month, start_col in month_cols.items():
        # Находим последний день месяца
        if month == 12:
            end_date_month = datetime(2026, 12, 31)
        else:
            end_date_month = datetime(2026, month + 1, 1) - timedelta(days=1)
        
        # Находим индекс последнего дня месяца
        end_col = start_col
        for col in range(start_col, col_index):
            col_letter = get_column_letter(col)
            if ws[f"{col_letter}1"].value and str(date.day) in ws[f"{col_letter}1"].value.split('\n')[0]:
                if date <= end_date_month:
                    end_col = col
        
        if end_col > start_col:
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            ws.merge_cells(f"{start_letter}2:{end_letter}2")
    
    # Третья строка: можно добавить рабочее время или другую информацию
    ws['A3'] = ""
    ws['B3'] = ""
    ws['C3'] = ""
    ws['D3'] = ""
    
    # Заполняем третью строку для календарных дней
    for col in range(5, col_index):
        col_letter = get_column_letter(col)
        cell = ws[f"{col_letter}3"]
        cell.value = ""  # Можно добавить "8ч" или другую информацию
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Устанавливаем ширину для первых столбцов
    ws.column_dimensions['A'].width = 4      # №
    ws.column_dimensions['B'].width = 25     # ФИО
    ws.column_dimensions['C'].width = 20     # Должность
    ws.column_dimensions['D'].width = 15     # Отдел
    
    # ========== ДОБАВЛЯЕМ СЛУЖЕБНУЮ ИНФОРМАЦИЮ ==========
    
    # Добавляем лист с легендой
    ws_legend = wb.create_sheet(title="Легенда")
    
    legend_data = [
        ["Обозначения в календаре:", ""],
        ["", ""],
        ["Цвет", "Тип дня", "Обозначение"],
        ["Белый", "Рабочий день", "Число + Пн/Вт/Ср/Чт/Пт"],
        ["Серый", "Выходной", "Число + Сб/Вс"],
        ["Красный", "Праздничный", "Число + ✶"],
        ["Желтый", "Предпраздничный", "Число + ◐"],
        ["", ""],
        ["Статистика 2026:", ""],
        ["Всего дней", len(all_dates)],
        ["Рабочих дней", len([d for d in all_dates if d.weekday() < 5 and d not in holidays])],
        ["Выходных", len([d for d in all_dates if d.weekday() >= 5])],
        ["Праздничных", len(holidays)],
    ]
    
    for i, row in enumerate(legend_data, 1):
        for j, value in enumerate(row, 1):
            ws_legend.cell(row=i, column=j, value=value)
    
    # Добавляем лист с праздниками
    ws_holidays = wb.create_sheet(title="Праздники")
    
    holidays_list = [
        ["Дата", "Праздник", "Тип дня"],
        ["01.01.2026", "Новый год", "Праздничный"],
        ["02.01.2026", "Новогодние каникулы", "Праздничный"],
        ["03.01.2026", "Новогодние каникулы", "Праздничный"],
        ["04.01.2026", "Новогодние каникулы", "Праздничный"],
        ["05.01.2026", "Новогодние каникулы", "Праздничный"],
        ["06.01.2026", "Новогодние каникулы", "Праздничный"],
        ["07.01.2026", "Рождество Христово", "Праздничный"],
        ["08.01.2026", "Новогодние каникулы", "Праздничный"],
        ["23.02.2026", "День защитника Отечества", "Праздничный"],
        ["08.03.2026", "Международный женский день", "Праздничный"],
        ["01.05.2026", "Праздник Весны и Труда", "Праздничный"],
        ["09.05.2026", "День Победы", "Праздничный"],
        ["12.06.2026", "День России", "Праздничный"],
        ["04.11.2026", "День народного единства", "Праздничный"],
    ]
    
    for i, row in enumerate(holidays_list, 1):
        for j, value in enumerate(row, 1):
            ws_holidays.cell(row=i, column=j, value=value)
    
    # Сохраняем файл
    output_file = "vacation_schedule_with_calendar_2026.xlsx"
    wb.save(output_file)
    
    print(f"✅ Файл создан: {output_file}")
    print(f"📊 Структура файла:")
    print(f"   1. Лист 'График отпусков 2026' - основной с календарем в шапке")
    print(f"   2. Лист 'Легенда' - расшифровка обозначений")
    print(f"   3. Лист 'Праздники' - список праздничных дней")
    print(f"\n📅 Календарь оформлен как заголовки столбцов:")
    print(f"   - Строка 1: число дня и обозначение (Пн/Вс/✶/◐)")
    print(f"   - Строка 2: месяц (объединен по всем дням месяца)")
    print(f"   - Строка 3: зарезервирована для дополнительной информации")
    print(f"\n🎨 Цветовая схема:")
    print(f"   ⬜ Белый - рабочие дни")
    print(f"   ⬜ Серый - выходные")
    print(f"   🟥 Красный - праздники")
    print(f"   🟨 Желтый - предпраздничные дни")
    
    # Показываем пример
    print(f"\n📋 Пример календаря (первые 10 дней января):")
    print("Колонка | Строка 1 | Строка 2")
    print("-" * 35)
    
    for col in range(5, 15):  # Первые 10 дней
        col_letter = get_column_letter(col)
        day_info = ws[f"{col_letter}1"].value or ""
        month_info = ws[f"{col_letter}2"].value or ""
        print(f"{col_letter:^7} | {day_info:^9} | {month_info:^9}")
    
    return output_file

# Запускаем создание календаря
if __name__ == "__main__":
    filename = create_calendar_header_2026()
    
    print(f"\n🚀 Инструкция по использованию:")
    print(f"1. Файл '{filename}' готов к использованию")
    print(f"2. В столбцах A-D добавьте данные сотрудников")
    print(f"3. В строках начиная с 4й отмечайте отпуска (например, закрашиванием ячеек)")
    print(f"4. Используйте фильтры для удобной работы с таблицей")
    
    print(f"\n💡 Советы:")
    print(f"• Используйте заливку для обозначения периодов отпуска")
    print(f"• Добавьте формулу для подсчета дней отпуска:")
    print(f"  =СЧЁТЕСЛИ(E4:ZZ4; \"✓\")  # где ✓ - отметка об отпуске")
    print(f"• Закрепите области: Вид → Закрепить области → Закрепить первые 3 строки")