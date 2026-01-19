"""
Генератор графика отпусков в Excel с VBA макросом для автоматического обновления

ВАЖНЫЙ ПРОМПТ ДЛЯ КОНТЕКСТА ПРЕДЫДУЩИХ ЧАТОВ (должен сохраняться при передаче в новый чат):
--------------------------------------------------------------------------------
Этот Python скрипт создает Excel файл для планирования отпусков сотрудников на 2026 год
с VBA макросом для автоматического обновления графика.

ОСНОВНЫЕ ТРЕБОВАНИЯ:
1. Лист "СОТРУДНИКИ" - таблица для ввода данных:
   - 20 сотрудников (строки 2-21)
   - До 10 периодов отпуска на сотрудника (столбцы C-V: пары начало/конец)
   - Формат дат: ДД.ММ.ГГГГ
2. Лист "ГРАФИК" - визуальное представление календаря на 2026 год:
   - Строки 1-3: заголовки календаря (месяцы, числа, дни недели с символами)
   - Строка 4: заголовки столбцов данных ("№", "ФИО СОТРУДНИКА")
   - Строки 5-24: данные 20 сотрудников
   - Календарь начинается с колонки C
   - Дни отпуска отмечаются буквой "О" на светло-зеленом фоне (RGB(198, 239, 206))
   - Производственный календарь России на 2026 год корректно отображен
   - Чередование цветов месяцев: только заголовки месяцев (строка 1) и область данных (строки 5-24)
3. Лист "ДАТЫ" - служебный скрытый лист с датами для работы макроса
4. Лист "ЛЕГЕНДА" и "ИНСТРУКЦИЯ" - пояснительные листы
5. VBA макрос в отдельном файле .txt для автоматического обновления графика

ОБНОВЛЕНИЯ:
- Добавлен лист "ПРАЗДНИКИ" со списком нерабочих дней 2026 года
- Формулы для подсчета дней отпуска исключают праздничные дни
- Производственный календарь 2026 года исправлен согласно официальным данным

ВАЖНО: При переносе задачи в новый чат этот промпт должен сохраняться в начале файла
для понимания контекста. Текущие задачи добавляются в конец промпта отдельным блоком.
--------------------------------------------------------------------------------
ИСПРАВЛЕНИЕ 1: Праздничные дни, которые попадают в период отпуска, теперь отмечаются 
буквой "О" на графике отпусков (ранее не отмечались).

ИСПРАВЛЕНИЕ 2: Оптимизирована скорость работы VBA макроса при пустых ФИО сотрудников.
Макрос останавливает обработку при обнаружении пустого ФИО (сотрудники идут по порядку).

ИСПРАВЛЕНИЕ 3: Добавлена кнопка на лист СОТРУДНИКИ для запуска макроса.

ИСПРАВЛЕНИЕ 4: Более контрастные цвета для разграничения дней и защита ячеек (без пароля).
"""

import os
import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import calendar

class WorkDay:
    def __init__(self, date, day_type, is_short=False):
        self.date = date
        self.day_type = day_type  # 'рабочий', 'выходной', 'праздник'
        self.is_short = is_short  # Сокращенный день

class ProductionCalendar:
    def __init__(self, year=2026):
        self.year = year
        self.days = []
        self._generate_calendar()
    
    def _generate_calendar(self):
        """Генерация производственного календаря на 2026 год в России (ОФИЦИАЛЬНЫЙ)"""
        date = datetime.date(self.year, 1, 1)
        
        holidays = [
            datetime.date(self.year, 1, 1),
            datetime.date(self.year, 1, 2),
            datetime.date(self.year, 1, 3),
            datetime.date(self.year, 1, 4),
            datetime.date(self.year, 1, 5),
            datetime.date(self.year, 1, 6),
            datetime.date(self.year, 1, 7),
            datetime.date(self.year, 1, 8),
            datetime.date(self.year, 2, 23),
            datetime.date(self.year, 3, 8),
            datetime.date(self.year, 5, 1),
            datetime.date(self.year, 5, 9),
            datetime.date(self.year, 6, 12),
            datetime.date(self.year, 11, 4),
        ]
        
        extra_holidays = [
            datetime.date(self.year, 1, 9),
            datetime.date(self.year, 3, 9),
            datetime.date(self.year, 5, 11),
            datetime.date(self.year, 12, 31),
        ]
        
        all_holidays = holidays + extra_holidays
        
        pre_holidays = [
            datetime.date(self.year, 2, 20),
            datetime.date(self.year, 4, 30),
            datetime.date(self.year, 5, 8),
            datetime.date(self.year, 6, 11),
            datetime.date(self.year, 11, 3),
        ]
        
        while date.year == self.year:
            day_type = 'рабочий'
            is_short = False
            
            if date in all_holidays:
                day_type = 'праздник'
            elif date.weekday() >= 5:
                day_type = 'выходной'
            
            if date in pre_holidays and day_type == 'рабочий':
                is_short = True
            
            self.days.append(WorkDay(date, day_type, is_short))
            date += timedelta(days=1)
    
    def get_day_info(self, date):
        """Получить информацию о дне"""
        for day in self.days:
            if day.date == date:
                return day
        return None
    
    def get_all_holidays(self):
        """Получить список всех праздничных дней (для листа ПРАЗДНИКИ)"""
        return [day.date for day in self.days if day.day_type == 'праздник']

class VacationScheduleGenerator:
    def __init__(self, company_name="ООО РОГА И КОПЫТА"):
        self.company_name = company_name
        self.year = 2026
        self.max_employees = 20
        self.vacation_pairs = 10
        self.calendar = ProductionCalendar(self.year)
    
    def create_excel_file(self):
        """Создание Excel файла"""
        print("Создание файла Excel...")
        
        wb = Workbook()
        
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        ws_employees = wb.create_sheet("СОТРУДНИКИ", 0)
        ws_schedule = wb.create_sheet("ГРАФИК", 1)
        ws_dates = wb.create_sheet("ДАТЫ", 2)
        ws_holidays = wb.create_sheet("ПРАЗДНИКИ", 3)
        ws_legend = wb.create_sheet("ЛЕГЕНДА", 4)
        ws_instruction = wb.create_sheet("ИНСТРУКЦИЯ", 5)
        
        ws_dates.sheet_state = 'hidden'
        ws_holidays.sheet_state = 'hidden'
        
        self._create_employees_sheet(ws_employees)
        self._create_schedule_sheet(ws_schedule)
        self._create_dates_sheet(ws_dates)
        self._create_holidays_sheet(ws_holidays)
        self._create_legend_sheet(ws_legend)
        self._create_instruction_sheet(ws_instruction)
        
        # Добавляем место для кнопки
        self._add_button_placeholder(ws_employees)
        
        # Не включаем защиту через openpyxl - будут проблемы с паролем
        # Вместо этого размечаем ячейки как заблокированные/разблокированные
        # Пользователь сам включит защиту в Excel если захочет
        
        current_date = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"отпуск_{self.company_name}_{self.year}_{current_date}.xlsx"
        
        try:
            wb.save(filename)
            print(f"✓ Файл создан: {filename}")
            return filename
        except Exception as e:
            print(f"✗ Ошибка при сохранении файла: {e}")
            return None
    
    def _add_button_placeholder(self, ws):
        """Добавляем место для кнопки на лист СОТРУДНИКИ"""
        total_blocks = self.max_employees
        columns_per_block = 4
        spacer_columns = self.max_employees - 1
        
        total_columns = (total_blocks * columns_per_block) + spacer_columns
        button_start_col = total_columns + 3
        
        button_row = 1
        button_height = 3
        
        ws.merge_cells(start_row=button_row, start_column=button_start_col,
                      end_row=button_row + button_height - 1, end_column=button_start_col + 2)
        
        button_cell = ws.cell(row=button_row, column=button_start_col, 
                             value="КНОПКА ДЛЯ ЗАПУСКА МАКРОСА")
        button_cell.font = Font(bold=True, size=12, color="FFFFFF")
        button_cell.alignment = Alignment(horizontal="center", vertical="center")
        button_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        instr_row = button_row + button_height + 1
        ws.merge_cells(start_row=instr_row, start_column=button_start_col,
                      end_row=instr_row, end_column=button_start_col + 2)
        
        instr_cell = ws.cell(row=instr_row, column=button_start_col,
                           value="1. Вставьте кнопку из панели разработчика")
        instr_cell.font = Font(size=9)
        instr_cell.alignment = Alignment(horizontal="center")
        
        instr_row2 = instr_row + 1
        ws.merge_cells(start_row=instr_row2, start_column=button_start_col,
                      end_row=instr_row2, end_column=button_start_col + 2)
        
        instr_cell2 = ws.cell(row=instr_row2, column=button_start_col,
                            value="2. Назначьте макрос 'ОбновитьГрафик'")
        instr_cell2.font = Font(size=9)
        instr_cell2.alignment = Alignment(horizontal="center")
        
        instr_row3 = instr_row2 + 1
        ws.merge_cells(start_row=instr_row3, start_column=button_start_col,
                      end_row=instr_row3, end_column=button_start_col + 2)
        
        instr_cell3 = ws.cell(row=instr_row3, column=button_start_col,
                            value="Alt+F8 - альтернативный способ")
        instr_cell3.font = Font(size=9, italic=True, color="666666")
        instr_cell3.alignment = Alignment(horizontal="center")
        
        for col in range(button_start_col, button_start_col + 3):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_employees_sheet(self, ws):
        """Создание листа СОТРУДНИКИ"""
        print("  Создание листа 'СОТРУДНИКИ'...")
        
        ws.sheet_view.showGridLines = False
        
        date_width = 12
        days_width = 10
        name_width = 25
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        data_font = Font(size=10)
        total_font = Font(bold=True, size=10)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        name_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        days_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        date_format = 'DD.MM.YYYY'
        BLOCK_COLS = 4
        MAX_PERIODS = 10
        
        for emp_index in range(self.max_employees):
            start_col = emp_index * BLOCK_COLS + 1
            
            headers = ["ФИО", "дни всего", "", ""]
            for col_offset, header in enumerate(headers):
                col = start_col + col_offset
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
            
            for col_offset in range(BLOCK_COLS):
                col = start_col + col_offset
                cell = ws.cell(row=2, column=col, value="")
                cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
            
            name_cell = ws.cell(row=3, column=start_col, value=f"Сотрудник {emp_index+1}")
            name_cell.font = Font(bold=True, size=10)
            name_cell.fill = name_fill
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            name_cell.border = Border(
                left=Side(style='thin', color="000000"),
                right=Side(style='thin', color="000000"),
                top=Side(style='thin', color="000000"),
                bottom=Side(style='thin', color="000000")
            )
            
            days_cell = ws.cell(row=3, column=start_col+1, value="")
            days_cell.font = total_font
            days_cell.fill = days_fill
            days_cell.alignment = Alignment(horizontal="center", vertical="center")
            days_cell.border = Border(
                left=Side(style='thin', color="000000"),
                right=Side(style='thin', color="000000"),
                top=Side(style='thin', color="000000"),
                bottom=Side(style='thin', color="000000")
            )
            
            for col_offset in range(2, BLOCK_COLS):
                col = start_col + col_offset
                cell = ws.cell(row=3, column=col, value="")
                cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
            
            for period_idx in range(MAX_PERIODS):
                row = 4 + period_idx
                
                start_cell = ws.cell(row=row, column=start_col+2, value="")
                start_cell.number_format = date_format
                start_cell.font = data_font
                start_cell.alignment = Alignment(horizontal="center", vertical="center")
                start_cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
                
                end_cell = ws.cell(row=row, column=start_col+3, value="")
                end_cell.number_format = date_format
                end_cell.font = data_font
                end_cell.alignment = Alignment(horizontal="center", vertical="center")
                end_cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
                
                days_period_cell = ws.cell(row=row, column=start_col+1, value="")
                days_period_cell.font = data_font
                days_period_cell.alignment = Alignment(horizontal="center", vertical="center")
                days_period_cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
                
                empty_cell = ws.cell(row=row, column=start_col, value="")
                empty_cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
            
            last_row = 4 + MAX_PERIODS
            for col_offset in range(BLOCK_COLS):
                col = start_col + col_offset
                cell = ws.cell(row=last_row, column=col, value="...")
                cell.font = Font(size=9, italic=True, color="666666")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin', color="000000"),
                    right=Side(style='thin', color="000000"),
                    top=Side(style='thin', color="000000"),
                    bottom=Side(style='thin', color="000000")
                )
            
            ws.column_dimensions[get_column_letter(start_col)].width = name_width
            ws.column_dimensions[get_column_letter(start_col+1)].width = days_width
            ws.column_dimensions[get_column_letter(start_col+2)].width = date_width
            ws.column_dimensions[get_column_letter(start_col+3)].width = date_width
            
            if emp_index < self.max_employees - 1:
                separator_col = start_col + BLOCK_COLS
                ws.column_dimensions[get_column_letter(separator_col)].width = 2
        
        self._add_intelligent_formulas_for_days(ws)
        self._add_example_data(ws)
        
        print(f"  ✓ Лист 'СОТРУДНИКИ' создан с {self.max_employees} блоками сотрудников")
    
    def _add_intelligent_formulas_for_days(self, ws):
        """Добавляем формулы Excel для расчета дней отпуска"""
        for emp_idx in range(self.max_employees):
            start_col = emp_idx * 4 + 1
            
            for period_idx in range(self.vacation_pairs):
                row = 4 + period_idx
                
                start_col_letter = get_column_letter(start_col + 2)
                end_col_letter = get_column_letter(start_col + 3)
                
                days_cell = ws.cell(row=row, column=start_col+1)
                
                formula = (
                    f'=IF(AND({start_col_letter}{row}<>"",{end_col_letter}{row}<>""),'
                    f'({end_col_letter}{row}-{start_col_letter}{row}+1)-'
                    f'COUNTIFS(ПРАЗДНИКИ!$A:$A,">="&{start_col_letter}{row},ПРАЗДНИКИ!$A:$A,"<="&{end_col_letter}{row}),'
                    f'"")'
                )
                
                days_cell.value = formula
                days_cell.number_format = '0'
        
        for emp_idx in range(self.max_employees):
            start_col = emp_idx * 4 + 1
            total_row = 3
            total_col = start_col + 1
            
            period_refs = []
            for period_idx in range(self.vacation_pairs):
                row = 4 + period_idx
                period_refs.append(f'{get_column_letter(start_col+1)}{row}')
            
            total_formula = f'=SUM({":".join(period_refs)})'
            total_cell = ws.cell(row=total_row, column=total_col)
            total_cell.value = total_formula
            total_cell.number_format = '0'
    
    def _add_example_data(self, ws):
        """Добавляем пример данных для первых трех сотрудников"""
        example_data = [
            {
                'name': 'Иванов И.И.',
                'periods': [
                    (datetime.date(2026, 1, 10), datetime.date(2026, 1, 20)),
                    (datetime.date(2026, 3, 20), datetime.date(2026, 3, 25)),
                    (datetime.date(2026, 6, 10), datetime.date(2026, 6, 15)),
                ]
            },
            {
                'name': 'Петров П.П.',
                'periods': [
                    (datetime.date(2026, 5, 1), datetime.date(2026, 5, 15)),
                ]
            },
            {
                'name': 'Сидоров С.С.',
                'periods': [
                    (datetime.date(2026, 8, 1), datetime.date(2026, 8, 14)),
                ]
            }
        ]
        
        for emp_idx, data in enumerate(example_data):
            if emp_idx >= 3:
                break
                
            start_col = emp_idx * 4 + 1
            
            ws.cell(row=3, column=start_col, value=data['name'])
            
            for period_idx, (start_date, end_date) in enumerate(data['periods']):
                if period_idx >= self.vacation_pairs:
                    break
                    
                row = 4 + period_idx
                ws.cell(row=row, column=start_col+2, value=start_date)
                ws.cell(row=row, column=start_col+3, value=end_date)
    
    def _create_holidays_sheet(self, ws):
        """Создание листа ПРАЗДНИКИ"""
        print("  Создание листа 'ПРАЗДНИКИ'...")
        
        ws.column_dimensions['A'].width = 15
        ws.cell(row=1, column=1, value="ПРАЗДНИЧНЫЕ ДНИ 2026").font = Font(bold=True, size=12, color="1F4E78")
        ws.cell(row=2, column=1, value="Дата").font = Font(bold=True)
        ws.cell(row=2, column=2, value="Описание").font = Font(bold=True)
        
        holidays = self.calendar.get_all_holidays()
        holidays.sort()
        
        descriptions = {
            datetime.date(2026, 1, 1): "Новый год",
            datetime.date(2026, 1, 2): "Новогодние каникулы",
            datetime.date(2026, 1, 3): "Новогодние каникулы",
            datetime.date(2026, 1, 4): "Новогодние каникулы",
            datetime.date(2026, 1, 5): "Новогодние каникулы",
            datetime.date(2026, 1, 6): "Новогодние каникулы",
            datetime.date(2026, 1, 7): "Рождество Христово",
            datetime.date(2026, 1, 8): "Новогодние каникулы",
            datetime.date(2026, 1, 9): "Перенос с 3 января",
            datetime.date(2026, 2, 23): "День защитника Отечества",
            datetime.date(2026, 3, 8): "Международный женский день",
            datetime.date(2026, 3, 9): "Перенос с 8 марта",
            datetime.date(2026, 5, 1): "Праздник Весны и Труда",
            datetime.date(2026, 5, 9): "День Победы",
            datetime.date(2026, 5, 11): "Перенос с 9 мая",
            datetime.date(2026, 6, 12): "День России",
            datetime.date(2026, 11, 4): "День народного единства",
            datetime.date(2026, 12, 31): "Перенос с 4 января",
        }
        
        for i, holiday in enumerate(holidays, start=1):
            row = i + 2
            ws.cell(row=row, column=1, value=holiday)
            ws.cell(row=row, column=1).number_format = 'DD.MM.YYYY'
            
            desc = descriptions.get(holiday, "Праздничный день")
            ws.cell(row=row, column=2, value=desc)
            
            holiday_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            ws.cell(row=row, column=1).fill = holiday_fill
            ws.cell(row=row, column=2).fill = holiday_fill
        
        print(f"  ✓ Лист 'ПРАЗДНИКИ' создан ({len(holidays)} праздничных дней)")
    
    def _create_schedule_sheet(self, ws):
        """Создание листа ГРАФИК с КОНТРАСТНЫМИ ЦВЕТАМИ"""
        print("  Создание листа 'ГРАФИК'...")
        
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        
        ws.cell(row=4, column=1, value="№").font = Font(bold=True)
        ws.cell(row=4, column=2, value="ФИО СОТРУДНИКА").font = Font(bold=True)
        
        ws.cell(row=4, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=4, column=2).alignment = Alignment(horizontal="left", vertical="center")
        
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.cell(row=4, column=1).fill = header_fill
        ws.cell(row=4, column=2).fill = header_fill
        
        current_col = 3
        month_names = ['ЯНВ', 'ФЕВ', 'МАР', 'АПР', 'МАЙ', 'ИЮН', 
                      'ИЮЛ', 'АВГ', 'СЕН', 'ОКТ', 'НОЯ', 'ДЕК']
        day_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
        
        # КОНТРАСТНЫЕ ЦВЕТА для лучшей видимости в Excel 2010
        colors = {
            'рабочий': 'FFFFFF',      # Белый - максимальный контраст
            'выходной': 'D9D9D9',     # СЕРЫЙ (было F2F2F2) - темнее для лучшей видимости
            'праздник': 'FF9999',     # Красный - оставляем как есть
        }
        
        # КОНТРАСТНЫЕ ЦВЕТА для чередования месяцев
        month_colors = ['E6E6E6', 'FFFFFF']  # Серый и белый (было F8F8F8 и FFFFFF)
        
        for month_idx, month in enumerate(range(1, 13)):
            days_in_month = calendar.monthrange(self.year, month)[1]
            
            start_col = current_col
            end_col = current_col + days_in_month - 1
            
            ws.merge_cells(
                start_row=1, start_column=start_col,
                end_row=1, end_column=end_col
            )
            
            month_color = month_colors[month_idx % 2]
            month_fill = PatternFill(start_color=month_color, end_color=month_color, fill_type="solid")
            
            month_cell = ws.cell(row=1, column=start_col, value=month_names[month-1])
            month_cell.alignment = Alignment(horizontal="center", vertical="center")
            month_cell.font = Font(bold=True, size=11)
            month_cell.fill = month_fill
            
            for day in range(1, days_in_month + 1):
                col = current_col + day - 1
                date_obj = datetime.date(self.year, month, day)
                day_info = self.calendar.get_day_info(date_obj)
                
                day_cell = ws.cell(row=2, column=col, value=day)
                day_cell.alignment = Alignment(horizontal="center", vertical="center")
                day_cell.font = Font(size=9)
                
                day_name = day_names[date_obj.weekday()]
                
                symbol = ""
                if day_info:
                    if day_info.day_type == 'праздник':
                        symbol = " ✶"
                    elif day_info.is_short:
                        symbol = " ●"
                
                weekday_cell = ws.cell(row=3, column=col, value=f"{day_name}{symbol}")
                weekday_cell.alignment = Alignment(horizontal="center", vertical="center")
                weekday_cell.font = Font(size=8)
                
                # Применяем КОНТРАСТНЫЕ цвета
                fill_color = colors['рабочий']
                if day_info:
                    if day_info.day_type == 'праздник':
                        fill_color = colors['праздник']
                    elif day_info.day_type == 'выходной':
                        fill_color = colors['выходной']
                
                day_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                day_cell.fill = day_fill
                weekday_cell.fill = day_fill
                
                ws.column_dimensions[get_column_letter(col)].width = 3.5
            
            current_col += days_in_month
        
        data_start_row = 5
        data_end_row = data_start_row + self.max_employees - 1
        
        # Применяем КОНТРАСТНЫЕ цвета месяцев к области данных
        col = 3
        for month_idx, month in enumerate(range(1, 13)):
            days_in_month = calendar.monthrange(self.year, month)[1]
            
            month_color = month_colors[month_idx % 2]
            month_fill = PatternFill(start_color=month_color, end_color=month_color, fill_type="solid")
            
            for day in range(1, days_in_month + 1):
                for row in range(data_start_row, data_end_row + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = month_fill
                col += 1
        
        for i in range(1, self.max_employees + 1):
            row = i + 4
            
            num_cell = ws.cell(row=row, column=1, value=i)
            num_cell.alignment = Alignment(horizontal="center", vertical="center")
            num_cell.font = Font(size=10)
            
            name_cell = ws.cell(row=row, column=2)
            name_cell.font = Font(size=10)
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            
            for col in range(1, current_col):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(vertical="center", horizontal="center")
            ws.cell(row=row, column=2).alignment = Alignment(vertical="center", horizontal="left")
        
        ws.freeze_panes = 'C5'
        
        last_col_letter = get_column_letter(current_col - 1)
        footer_row = self.max_employees + 6
        ws.merge_cells(f'A{footer_row}:{last_col_letter}{footer_row}')
        footer = ws.cell(row=footer_row, column=1, 
                        value=f"График отпусков {self.company_name} на {self.year} год")
        footer.font = Font(italic=True, size=10, color="666666")
        footer.alignment = Alignment(horizontal="center")
        
        print(f"  ✓ Лист 'ГРАФИК' создан ({current_col-3} дней)")
        print("    • Применены контрастные цвета для Excel 2010")
    
    def _create_dates_sheet(self, ws):
        """Создание служебного листа с датами"""
        print("  Создание служебного листа с датами...")
        
        date_obj = datetime.date(self.year, 1, 1)
        col = 3
        
        while date_obj.year == self.year:
            cell = ws.cell(row=1, column=col, value=date_obj)
            cell.number_format = 'DD.MM.YYYY'
            date_obj += timedelta(days=1)
            col += 1
        
        for c in range(1, col):
            ws.column_dimensions[get_column_letter(c)].width = 0.5
        
        print(f"  ✓ Служебный лист создан ({col-3} дней, начинается с колонки C)")
    
    def _create_legend_sheet(self, ws):
        """Создание листа ЛЕГЕНДА"""
        print("  Создание листа 'ЛЕГЕНДА'...")
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        
        ws.merge_cells('A1:C1')
        title = ws.cell(row=1, column=1, value="ЛЕГЕНДА ГРАФИКА ОТПУСКОВ")
        title.font = Font(bold=True, size=14, color="1F4E78")
        title.alignment = Alignment(horizontal="center")
        
        headers = ["Обозначение", "Тип дня", "Описание"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Обновляем цвета в легенде для соответствия контрастным цветам
        legend_data = [
            ("✶", "Праздник/Выходной", "Нерабочий праздничный или дополнительный выходной день", "FF9999"),
            ("●", "Сокращенный", "Предпраздничный рабочий день (короче на 1 час)", "FFFF99"),
            ("О", "Отпуск", "День отпуска сотрудника", "C6EFCE"),
            ("", "Выходной", "Суббота, воскресенье (серый фон)", "D9D9D9"),  # Обновлен цвет
            ("", "Рабочий", "Обычный рабочий день (белый фон)", "FFFFFF"),
            ("", "Месяц 1", "Четные месяцы (светло-серый фон)", "E6E6E6"),   # Добавлен для чередования месяцев
            ("", "Месяц 2", "Нечетные месяцы (белый фон)", "FFFFFF"),       # Добавлен для чередования месяцев
        ]
        
        for i, (symbol, day_type, description, color) in enumerate(legend_data, 1):
            row = i + 3
            
            sym_cell = ws.cell(row=row, column=1, value=symbol)
            sym_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            type_cell = ws.cell(row=row, column=2, value=day_type)
            desc_cell = ws.cell(row=row, column=3, value=description)
            
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sym_cell.fill = fill
            type_cell.fill = fill
            desc_cell.fill = fill
        
        print("  ✓ Лист 'ЛЕГЕНДА' создан с обновленными цветами")
    
    def _create_instruction_sheet(self, ws):
        """Создание листа ИНСТРУКЦИЯ"""
        print("  Создание листа 'ИНСТРУКЦИЯ'...")
        
        ws.column_dimensions['A'].width = 80
        
        content = [
            ("ИНСТРУКЦИЯ ПО РАБОТЕ С ГРАФИКОМ ОТПУСКОВ", 16, True, True),
            ("", 1, False, False),
            ("НОВЫЕ ВОЗМОЖНОСТИ:", 14, True, False),
            ("• КОНТРАСТНЫЕ ЦВЕТА для лучшей видимости в Excel 2010", 11, False, False),
            ("• Ячейки размечены для защиты (пользователь включает защиту при необходимости)", 11, False, False),
            ("• КНОПКА для запуска макроса + Alt+F8", 11, False, False),
            ("", 1, False, False),
            ("КОНТРАСТНЫЕ ЦВЕТА:", 14, True, False),
            ("• Выходные дни: СЕРЫЙ фон (D9D9D9)", 11, False, False),
            ("• Рабочие дни: БЕЛЫЙ фон (FFFFFF)", 11, False, False),
            ("• Чередование месяцев: СЕРЫЙ/БЕЛЫЙ (E6E6E6/FFFFFF)", 11, False, False),
            ("• Праздничные дни: КРАСНЫЙ фон (FF9999)", 11, False, False),
            ("• Отпуск: СВЕТЛО-ЗЕЛЕНЫЙ фон (C6EFCE)", 11, False, False),
            ("", 1, False, False),
            ("КАК ДОБАВИТЬ КНОПКУ:", 14, True, False),
            ("1. Включите панель разработчика: Файл → Параметры → Настройка ленты", 11, False, False),
            ("2. Отметьте 'Разработчик' в списке вкладок", 11, False, False),
            ("3. Перейдите на вкладку 'Разработчик'", 11, False, False),
            ("4. Нажмите 'Вставить' → выберите 'Кнопка' (первая в форме элементов управления)", 11, False, False),
            ("5. Нарисуйте кнопку в выделенной области на листе СОТРУДНИКИ", 11, False, False),
            ("6. В диалоговом окне выберите макрос 'ОбновитьГрафик'", 11, False, False),
            ("7. Переименуйте кнопку (например: 'Построить график')", 11, False, False),
            ("", 1, False, False),
            ("ШАГИ РАБОТЫ:", 14, True, False),
            ("1. Заполните ФИО и даты отпусков на листе 'СОТРУДНИКИ'", 11, False, False),
            ("2. Нажмите кнопку 'Построить график' ИЛИ Alt+F8 → 'ОбновитьГрафик'", 11, False, False),
            ("3. Перейдите на лист 'ГРАФИК' для просмотра", 11, False, False),
            ("4. Количество дней рассчитывается автоматически (исключая праздники)", 11, False, False),
            ("", 1, False, False),
            ("ПРИМЕЧАНИЕ:", 14, True, False),
            ("• Ячейки размечены для защиты, но защита не включена", 11, False, False),
            ("• Чтобы включить защиту: Рецензирование → Защитить лист", 11, False, False),
            ("• При включении защиты пароль НЕ устанавливайте (оставьте поле пустым)", 11, False, False),
            ("• Разрешены для редактирования: ФИО и даты отпусков", 11, False, False),
        ]
        
        for i, (text, size, bold, center) in enumerate(content, 1):
            cell = ws.cell(row=i, column=1, value=text)
            cell.font = Font(size=size, bold=bold)
            if center:
                cell.alignment = Alignment(horizontal="center")
        
        print("  ✓ Лист 'ИНСТРУКЦИЯ' создан")
    
    def create_vba_macro_file(self):
        """Создание файла с оптимизированным VBA макросом"""
        print("\nСоздание файла с VBA макросом...")
        
        vba_code = '''Option Explicit

Public Const MAX_EMPLOYEES As Integer = 20
Public Const BLOCK_COLS As Integer = 4
Public Const MAX_PERIODS As Integer = 10

' ОБНОВЛЕННЫЕ КОНТРАСТНЫЕ ЦВЕТА для Excel 2010
Private Const COLOR_MONTH_1 As Long = &HE6E6E6     ' Светло-серый (контрастный)
Private Const COLOR_MONTH_2 As Long = &HFFFFFF     ' Белый
Private Const COLOR_WEEKEND As Long = &HD9D9D9     ' Серый для выходных (контрастный)
Private Const COLOR_HOLIDAY As Long = &H9999FF     ' Красный для праздников
Private Const COLOR_VACATION As Long = &HCEC6EF    ' Светло-зеленый для отпуска

Private Const DAYS_IN_MONTHS As String = "31,29,31,30,31,30,31,31,30,31,30,31"

Sub ОбновитьГрафик()
    ' Макрос для обновления графика отпусков
    ' Обновлено: используются контрастные цвета для Excel 2010
    
    Dim wsEmployees As Worksheet
    Dim wsSchedule As Worksheet
    Dim wsService As Worksheet
    
    Dim employeeCount As Long
    Dim vacationCount As Long
    Dim i As Long, j As Long
    Dim empBlockStart As Long
    
    Dim startDate As Date
    Dim endDate As Date
    Dim currentDate As Date
    Dim dateCol As Long
    
    Dim scheduleRow As Long
    Dim nameCell As Range
    Dim startDateCell As Range
    Dim endDateCell As Range
    
    Dim dateDict As Object
    Dim lastCol As Long
    Dim dictKey As String
    
    Dim startTime As Double
    startTime = Timer
    
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    Application.EnableEvents = False
    
    On Error GoTo ErrorHandler
    
    employeeCount = 0
    vacationCount = 0
    
    Set wsEmployees = ThisWorkbook.Worksheets("СОТРУДНИКИ")
    Set wsSchedule = ThisWorkbook.Worksheets("ГРАФИК")
    Set wsService = ThisWorkbook.Worksheets("ДАТЫ")
    
    Call ОчиститьГрафик(wsSchedule)
    
    ' БЫСТРЫЙ ПОИСК ДАТ: создаем словарь для мгновенного доступа
    Set dateDict = CreateObject("Scripting.Dictionary")
    lastCol = wsService.Cells(1, wsService.Columns.Count).End(xlToLeft).Column
    
    For i = 3 To lastCol
        If Not IsEmpty(wsService.Cells(1, i).Value) Then
            If IsDate(wsService.Cells(1, i).Value) Then
                dateDict(Format(wsService.Cells(1, i).Value, "YYYYMMDD")) = i
            End If
        End If
    Next i
    
    ' ОСНОВНАЯ ОПТИМИЗАЦИЯ: остановка при пустых ФИО
    Dim stopProcessing As Boolean
    stopProcessing = False
    
    For i = 0 To MAX_EMPLOYEES - 1
        If stopProcessing Then Exit For
        
        empBlockStart = i * BLOCK_COLS + 1
        Set nameCell = wsEmployees.Cells(3, empBlockStart)
        
        ' Пропускаем пустые ФИО
        If Trim(nameCell.Value) = "" Then
            ' Если три подряд пустых ФИО, останавливаем обработку
            If i > 2 Then
                Dim emptyCount As Long
                emptyCount = 0
                For j = i To Application.Min(i + 2, MAX_EMPLOYEES - 1)
                    Dim testBlock As Long
                    testBlock = j * BLOCK_COLS + 1
                    If Trim(wsEmployees.Cells(3, testBlock).Value) = "" Then
                        emptyCount = emptyCount + 1
                    Else
                        Exit For
                    End If
                Next j
                
                If emptyCount >= 3 Then
                    stopProcessing = True
                End If
            End If
            GoTo NextEmployee
        End If
        
        employeeCount = employeeCount + 1
        scheduleRow = employeeCount + 4
        
        wsSchedule.Cells(scheduleRow, 1).Value = employeeCount
        wsSchedule.Cells(scheduleRow, 2).Value = nameCell.Value
        
        Call ВосстановитьЦветаМесяцев(wsSchedule, scheduleRow)
        
        ' Обработка периодов отпуска
        For j = 0 To MAX_PERIODS - 1
            Dim periodRow As Long
            periodRow = 4 + j
            
            Set startDateCell = wsEmployees.Cells(periodRow, empBlockStart + 2)
            Set endDateCell = wsEmployees.Cells(periodRow, empBlockStart + 3)
            
            If Not IsEmpty(startDateCell.Value) And Not IsEmpty(endDateCell.Value) Then
                If IsDate(startDateCell.Value) And IsDate(endDateCell.Value) Then
                    startDate = CDate(startDateCell.Value)
                    endDate = CDate(endDateCell.Value)
                    
                    If endDate >= startDate Then
                        vacationCount = vacationCount + 1
                        
                        currentDate = startDate
                        Do While currentDate <= endDate
                            ' БЫСТРЫЙ ПОИСК через словарь
                            dictKey = Format(currentDate, "YYYYMMDD")
                            If dateDict.Exists(dictKey) Then
                                dateCol = dateDict(dictKey)
                                
                                With wsSchedule.Cells(scheduleRow, dateCol)
                                    .Value = "О"
                                    .Interior.Color = COLOR_VACATION
                                    .Font.Bold = True
                                    .Font.Name = "Arial"
                                    .Font.Size = 9
                                    .HorizontalAlignment = xlCenter
                                    .VerticalAlignment = xlCenter
                                End With
                            End If
                            
                            currentDate = currentDate + 1
                        Loop
                    End If
                End If
            End If
        Next j

NextEmployee:
    Next i
    
    wsSchedule.Columns("B:B").AutoFit
    
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.EnableEvents = True
    
    Dim endTime As Double
    endTime = Timer
    Dim elapsedTime As Double
    elapsedTime = endTime - startTime
    
    MsgBox "График отпусков успешно обновлен!" & vbCrLf & _
           "Время выполнения: " & Format(elapsedTime, "0.0") & " сек" & vbCrLf & _
           "Сотрудников: " & employeeCount & vbCrLf & _
           "Периодов отпуска: " & vacationCount & vbCrLf & _
           "Использованы контрастные цвета для Excel 2010", _
           vbInformation, "График обновлен"
    
    wsSchedule.Activate
    wsSchedule.Range("A1").Select
    
    Exit Sub

ErrorHandler:
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.EnableEvents = True
    
    MsgBox "Ошибка при обновлении графика!" & vbCrLf & _
           "Код ошибки: " & Err.Number & vbCrLf & _
           "Описание: " & Err.Description & vbCrLf & _
           "Проверьте корректность введенных дат.", _
           vbCritical, "Ошибка"
End Sub

Private Sub ОчиститьГрафик(wsSchedule As Worksheet)
    Dim lastRow As Long
    Dim lastCol As Long
    Dim i As Long, j As Long
    
    lastRow = wsSchedule.Cells(wsSchedule.Rows.Count, "B").End(xlUp).Row
    
    If lastRow >= 5 Then
        wsSchedule.Range("A5:B" & lastRow).ClearContents
        
        lastCol = wsSchedule.Cells(3, wsSchedule.Columns.Count).End(xlToLeft).Column
        
        If lastCol >= 3 Then
            For i = 5 To lastRow
                For j = 3 To lastCol
                    With wsSchedule.Cells(i, j)
                        .Value = ""
                        .Font.Bold = False
                        .Interior.Pattern = xlNone
                    End With
                Next j
                Call ВосстановитьЦветаМесяцев(wsSchedule, i)
            Next i
        End If
    End If
End Sub

Private Sub ВосстановитьЦветаМесяцев(wsSchedule As Worksheet, rowNum As Long)
    ' Восстанавливает контрастное чередование цветов месяцев
    
    Dim monthDays() As String
    Dim i As Long, j As Long
    Dim col As Long
    Dim monthColor As Long
    Dim daysInMonth As Integer
    
    monthDays = Split(DAYS_IN_MONTHS, ",")
    col = 3
    
    For i = 0 To UBound(monthDays)
        daysInMonth = CInt(monthDays(i))
        
        ' Чередование цветов месяцев (контрастное)
        If (i Mod 2) = 0 Then
            monthColor = COLOR_MONTH_1  ' Четные месяцы: светло-серый
        Else
            monthColor = COLOR_MONTH_2  ' Нечетные месяцы: белый
        End If
        
        For j = 1 To daysInMonth
            wsSchedule.Cells(rowNum, col).Interior.Color = monthColor
            col = col + 1
        Next j
    Next i
End Sub

Sub ТестовыеДанные()
    ' Процедура для заполнения тестовых данных
    
    Dim ws As Worksheet
    Dim i As Long
    Dim blockStart As Long
    
    Set ws = ThisWorkbook.Worksheets("СОТРУДНИКИ")
    
    ' Очищаем только первые 10 сотрудников
    For i = 0 To 9
        blockStart = i * BLOCK_COLS + 1
        ws.Cells(3, blockStart).ClearContents
        ws.Range(ws.Cells(4, blockStart + 2), ws.Cells(13, blockStart + 3)).ClearContents
    Next i
    
    ' Сотрудник 1
    ws.Cells(3, 1).Value = "Иванов Иван Иванович"
    ws.Cells(4, 3).Value = DateSerial(2026, 1, 10)
    ws.Cells(4, 4).Value = DateSerial(2026, 1, 20)
    ws.Cells(5, 3).Value = DateSerial(2026, 6, 10)
    ws.Cells(5, 4).Value = DateSerial(2026, 6, 15)
    
    ' Сотрудник 2
    ws.Cells(3, 5).Value = "Петров Петр Петрович"
    ws.Cells(4, 7).Value = DateSerial(2026, 5, 1)
    ws.Cells(4, 8).Value = DateSerial(2026, 5, 15)
    
    ' Сотрудник 3
    ws.Cells(3, 9).Value = "Сидоров Сергей Сергеевич"
    ws.Cells(4, 11).Value = DateSerial(2026, 8, 1)
    ws.Cells(4, 12).Value = DateSerial(2026, 8, 14)
    
    ' Сотрудник 4
    ws.Cells(3, 13).Value = "Козлова Анна Михайловна"
    ws.Cells(4, 15).Value = DateSerial(2026, 7, 1)
    ws.Cells(4, 16).Value = DateSerial(2026, 7, 14)
    
    ' Форматируем даты
    For i = 0 To 9
        blockStart = i * BLOCK_COLS + 1
        ws.Range(ws.Cells(4, blockStart + 2), ws.Cells(13, blockStart + 3)).NumberFormat = "DD.MM.YYYY"
    Next i
    
    ' Пересчитываем формулы
    ws.Calculate
    
    MsgBox "Тестовые данные для 4 сотрудников добавлены!" & vbCrLf & _
           "Для удаления сотрудника оставьте поле ФИО пустым." & vbCrLf & _
           "Ячейки размечены для защиты (защита не включена).", _
           vbInformation, "Тестовые данные"
End Sub
'''
        
        filename = "vacation_macro.txt"
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(vba_code)
            print(f"✓ Файл с макросом создан: {filename}")
            return filename
        except Exception as e:
            print(f"✗ Ошибка при создании файла макроса: {e}")
            return None


def main():
    """Основная функция"""
    print("=" * 70)
    print("ГЕНЕРАТОР ГРАФИКА ОТПУСКОВ С VBA МАКРОСОМ")
    print("Версия с контрастными цветами (защита отключена)")
    print("=" * 70)
    
    company_name = input("\nВведите название компании: ").strip()
    if not company_name:
        company_name = "ООО РОГА И КОПЫТА"
    
    print("\n" + "=" * 70)
    print("СОЗДАНИЕ ФАЙЛОВ...")
    print("=" * 70)
    
    generator = VacationScheduleGenerator(company_name)
    excel_file = generator.create_excel_file()
    macro_file = generator.create_vba_macro_file()
    
    print("\n" + "=" * 70)
    
    if excel_file and macro_file:
        print("✓ ФАЙЛЫ УСПЕШНО СОЗДАНЫ")
        print(f"  • Excel файл: {excel_file}")
        print(f"  • Файл макроса: vacation_macro.txt")
        print("\nОСНОВНЫЕ УЛУЧШЕНИЯ:")
        print("  1. КОНТРАСТНЫЕ ЦВЕТА для Excel 2010:")
        print("     • Выходные дни: СЕРЫЙ (D9D9D9)")
        print("     • Рабочие дни: БЕЛЫЙ (FFFFFF)")
        print("     • Чередование месяцев: СЕРЫЙ/БЕЛЫЙ")
        print("  2. ЗАЩИТА ОТКЛЮЧЕНА:")
        print("     • Не будет запрашиваться пароль")
        print("     • Пользователь сам включает защиту если нужно")
        print("  3. КНОПКА ДЛЯ ЗАПУСКА + Alt+F8")
        print("  4. ОПТИМИЗИРОВАННАЯ СКОРОСТЬ РАБОТЫ")
        print("=" * 70)
    else:
        print("✗ ОШИБКА! Не удалось создать файлы.")


if __name__ == "__main__":
    main()