#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
СОЗДАТЕЛЬ ГРАФИКА ОТПУСКОВ 2026
Создает чистый Excel файл для последующего добавления макроса VBA
"""

import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_russian_calendar_2026():
    """Возвращает производственный календарь России на 2026 год"""
    holidays = [
        (2026, 1, 1), (2026, 1, 2), (2026, 1, 3), (2026, 1, 4),
        (2026, 1, 5), (2026, 1, 6), (2026, 1, 7), (2026, 1, 8),
        (2026, 1, 9), (2026, 2, 23), (2026, 3, 8), (2026, 5, 1),
        (2026, 5, 9), (2026, 6, 12), (2026, 11, 4),
    ]
    
    pre_holidays = [
        (2026, 2, 20), (2026, 3, 7), (2026, 5, 8),
        (2026, 6, 11), (2026, 11, 3), (2026, 12, 31),
    ]
    
    working_saturdays = [
        (2026, 2, 21), (2026, 11, 14),
    ]
    
    calendar = {}
    start_date = datetime(2026, 1, 1)
    
    for i in range(366):
        current_date = start_date + timedelta(days=i)
        date_key = current_date.date()
        weekday = current_date.weekday()
        
        is_holiday = (current_date.year, current_date.month, current_date.day) in holidays
        is_pre_holiday = (current_date.year, current_date.month, current_date.day) in pre_holidays
        is_working_saturday = (current_date.year, current_date.month, current_date.day) in working_saturdays
        
        if is_holiday:
            day_type = "holiday"
        elif is_pre_holiday:
            day_type = "pre_holiday"
        elif is_working_saturday:
            day_type = "work_saturday"
        elif weekday >= 5:
            day_type = "weekend"
        else:
            day_type = "workday"
        
        calendar[date_key] = {
            'date': current_date,
            'day': current_date.day,
            'month': current_date.month,
            'weekday': weekday,
            'day_type': day_type,
        }
    
    return calendar

def create_calendar_sheet(ws, calendar):
    """Создает календарь на листе"""
    # Стили для месяцев
    month_colors = {
        1: "4F81BD", 2: "8064A2", 3: "9BBB59", 4: "C0504D",
        5: "F79646", 6: "1F497D", 7: "948A54", 8: "31869B",
        9: "E26B0A", 10: "60497A", 11: "C00000", 12: "366092"
    }
    
    # Группируем дни по месяцам
    months = {}
    for date_info in calendar.values():
        month = date_info['month']
        if month not in months:
            months[month] = []
        months[month].append(date_info)
    
    # Сортируем месяцы
    sorted_months = sorted(months.keys())
    
    current_col = 3  # Начинаем с колонки C
    month_names = {
        1: "ЯНВ", 2: "ФЕВ", 3: "МАР", 4: "АПР",
        5: "МАЙ", 6: "ИЮН", 7: "ИЮЛ", 8: "АВГ",
        9: "СЕН", 10: "ОКТ", 11: "НОЯ", 12: "ДЕК"
    }
    
    weekday_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    for month_num in sorted_months:
        month_days = months[month_num]
        start_col = current_col
        end_col = current_col + len(month_days) - 1
        
        # Объединяем для названия месяца
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        ws.merge_cells(f"{start_letter}1:{end_letter}1")
        
        # Название месяца
        month_cell = ws[f"{start_letter}1"]
        month_cell.value = month_names[month_num]
        month_cell.fill = PatternFill(start_color=month_colors[month_num], fill_type="solid")
        month_cell.font = Font(color="FFFFFF", bold=True)
        month_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Дни месяца
        for i, day_info in enumerate(month_days):
            col = current_col + i
            
            # Число дня
            day_cell = ws.cell(row=2, column=col, value=day_info['day'])
            day_cell.alignment = Alignment(horizontal="center")
            day_cell.font = Font(bold=True, size=9)
            
            # День недели
            weekday = weekday_names[day_info['weekday']]
            
            # Символы для особых дней
            symbol = ""
            bg_color = "FFFFFF"
            if day_info['day_type'] == 'holiday':
                symbol = " ✶"
                bg_color = "FF9999"
            elif day_info['day_type'] == 'pre_holiday':
                symbol = " ◐"
                bg_color = "FFFF99"
            elif day_info['day_type'] == 'work_saturday':
                symbol = " ⚒"
                bg_color = "CCFFCC"
            elif day_info['day_type'] == 'weekend':
                bg_color = "E6E6E6"
            
            day_name_cell = ws.cell(row=3, column=col, value=f"{weekday}{symbol}")
            day_name_cell.alignment = Alignment(horizontal="center")
            day_name_cell.font = Font(size=9)
            day_name_cell.fill = PatternFill(start_color=bg_color, fill_type="solid")
            
            # Устанавливаем ширину
            ws.column_dimensions[get_column_letter(col)].width = 4.5
        
        current_col += len(month_days)
    
    return current_col - 1  # Возвращаем последний столбец

def create_vacation_file():
    """Создает Excel файл с графиком отпусков"""
    
    print("=" * 70)
    print("СОЗДАТЕЛЬ ГРАФИКА ОТПУСКОВ 2026")
    print("=" * 70)
    
    # Генерируем календарь
    print("\n📅 Генерирую календарь...")
    calendar = get_russian_calendar_2026()
    
    # Создаем книгу Excel
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 1. ЛИСТ "ДАННЫЕ" (вместо "СОТРУДНИКИ")
    print("📝 Создаю лист с данными...")
    ws_data = wb.create_sheet(title="ДАННЫЕ")
    
    # Заголовки
    headers = ["ID", "ФИО", "Отпуск1_начало", "Отпуск1_конец", "Отпуск1_дни",
               "Отпуск2_начало", "Отпуск2_конец", "Отпуск2_дни",
               "Отпуск3_начало", "Отпуск3_конец", "Отпуск3_дни"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(start_color="1F497D", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Пример данных
    employees = [
        ["1", "ИВАНОВ ИВАН ИВАНОВИЧ", "10.01.2026", "25.01.2026", "", 
         "15.07.2026", "01.08.2026", "", "", "", ""],
        ["2", "ПЕТРОВ ПЕТР ПЕТРОВИЧ", "15.02.2026", "25.02.2026", "",
         "01.09.2026", "14.09.2026", "", "", "", ""],
        ["3", "СИДОРОВА МАРИЯ ВЛАДИМИРОВНА", "01.03.2026", "14.03.2026", "",
         "10.10.2026", "20.10.2026", "", "", "", ""],
        ["4", "КОЗЛОВ АЛЕКСЕЙ НИКОЛАЕВИЧ", "01.04.2026", "10.04.2026", "",
         "01.11.2026", "10.11.2026", "", "", "", ""],
        ["5", "МОРОЗОВА ЕЛЕНА СЕРГЕЕВНА", "10.05.2026", "24.05.2026", "",
         "15.12.2026", "31.12.2026", "", "", "", ""],
    ]
    
    for row_idx, emp_data in enumerate(employees, start=2):
        for col_idx, value in enumerate(emp_data, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Закрашиваем через строку
        if row_idx % 2 == 0:
            for col in range(1, 12):
                ws_data.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="F2F2F2", fill_type="solid")
    
    # Настраиваем ширину столбцов
    widths = [5, 30, 12, 12, 8, 12, 12, 8, 12, 12, 8]
    for i, width in enumerate(widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = width
    
    # 2. ЛИСТ "ГРАФИК" (пустой, будет заполняться макросом)
    print("📊 Создаю лист графика...")
    ws_graph = wb.create_sheet(title="ГРАФИК")
    
    # Заголовки графика
    ws_graph['A1'] = "№"
    ws_graph['B1'] = "ФИО"
    
    for col in [1, 2]:
        cell = ws_graph.cell(row=1, column=col)
        cell.fill = PatternFill(start_color="1F497D", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws_graph.column_dimensions['A'].width = 5
    ws_graph.column_dimensions['B'].width = 30
    
    # Добавляем сотрудников
    for i, emp in enumerate(employees, start=2):
        ws_graph.cell(row=i, column=1, value=emp[0]).alignment = Alignment(horizontal="center")
        ws_graph.cell(row=i, column=2, value=emp[1])
        
        # Границы
        for col in [1, 2]:
            ws_graph.cell(row=i, column=col).border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Закрашивание
        if i % 2 == 0:
            for col in [1, 2]:
                ws_graph.cell(row=i, column=col).fill = PatternFill(
                    start_color="F8F8F8", fill_type="solid")
    
    # Создаем календарь на графике
    last_col = create_calendar_sheet(ws_graph, calendar)
    
    # 3. ЛИСТ "ИНСТРУКЦИЯ"
    print("📋 Создаю инструкцию...")
    ws_help = wb.create_sheet(title="ИНСТРУКЦИЯ")
    
    instructions = [
        ["ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ"],
        [""],
        ["1. ДОБАВЛЕНИЕ МАКРОСА:"],
        ["   - Откройте файл в Excel"],
        ["   - Нажмите Alt+F11 (откроется редактор VBA)"],
        ["   - В меню выберите: Insert → Module"],
        ["   - Скопируйте код макроса из файла 'vacation_macro.txt'"],
        ["   - Закройте редактор VBA (Ctrl+Q)"],
        [""],
        ["2. ЗАПУСК МАКРОСА:"],
        ["   - Вернитесь в Excel"],
        ["   - Нажмите Alt+F8"],
        ["   - Выберите макрос 'UpdateSchedule'"],
        ["   - Нажмите 'Выполнить'"],
        [""],
        ["3. РАБОТА С ДАННЫМИ:"],
        ["   - Вносите даты отпусков на листе 'ДАННЫЕ'"],
        ["   - Формат дат: ДД.ММ.ГГГГ"],
        ["   - После изменения дат запускайте макрос"],
        ["   - График автоматически обновится на листе 'ГРАФИК'"],
        [""],
        ["4. ДОБАВЛЕНИЕ СОТРУДНИКОВ:"],
        ["   - Добавляйте новые строки в конец на листе 'ДАННЫЕ'"],
        ["   - Формат должен совпадать с существующими строками"],
        ["   - ID должен быть уникальным"],
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        ws_help.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx in [1, 3, 10, 17]:
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    
    ws_help.column_dimensions['A'].width = 60
    
    # Сохраняем файл
    filename = "отпуск_график_готовый.xlsx"
    print(f"\n💾 Сохраняю файл: {filename}")
    wb.save(filename)
    
    print("\n" + "=" * 70)
    print("✅ EXCEL ФАЙЛ УСПЕШНО СОЗДАН!")
    print("=" * 70)
    
    # Создаем файл с макросом
    create_macro_file()
    
    return filename

def create_macro_file():
    """Создает отдельный файл с кодом макроса"""
    macro_code = '''Attribute VB_Name = "Module1"
' МАКРОС ДЛЯ ОБНОВЛЕНИЯ ГРАФИКА ОТПУСКОВ
' Автор: Генератор графиков отпусков

Option Explicit

' ОСНОВНОЙ МАКРОС - запускайте эту процедуру
Public Sub UpdateSchedule()
    Dim wsData As Worksheet    ' Лист с данными
    Dim wsGraph As Worksheet   ' Лист с графиком
    Dim lastRow As Long        ' Последняя строка с данными
    Dim lastCol As Long        ' Последний столбец в графике
    Dim i As Long, j As Long   ' Счетчики
    Dim startDate As Date      ' Начало отпуска
    Dim endDate As Date        ' Конец отпуска
    Dim currentDate As Date    ' Текущая дата в цикле
    Dim colNum As Long         ' Номер столбца для даты
    Dim daysCount As Long      ' Количество дней отпуска
    
    ' Отключаем обновление экрана для скорости
    Application.ScreenUpdating = False
    Application.Calculation = xlCalculationManual
    Application.DisplayAlerts = False
    
    On Error GoTo ErrorHandler
    
    ' Находим наши листы
    Set wsData = ThisWorkbook.Worksheets("ДАННЫЕ")
    Set wsGraph = ThisWorkbook.Worksheets("ГРАФИК")
    
    ' 1. ОЧИСТКА СТАРОГО ГРАФИКА
    Call ClearOldSchedule(wsGraph)
    
    ' 2. ОБНОВЛЕНИЕ ДАННЫХ СОТРУДНИКОВ
    lastRow = wsData.Cells(wsData.Rows.Count, 1).End(xlUp).Row
    
    For i = 2 To lastRow
        If wsData.Cells(i, 2).Value <> "" Then ' Если есть ФИО
            ' ОБРАБОТКА ПЕРВОГО ОТПУСКА
            Call ProcessVacation(wsData, wsGraph, i, 3, 4, 5)
            
            ' ОБРАБОТКА ВТОРОГО ОТПУСКА
            Call ProcessVacation(wsData, wsGraph, i, 6, 7, 8)
            
            ' ОБРАБОТКА ТРЕТЬЕГО ОТПУСКА
            Call ProcessVacation(wsData, wsGraph, i, 9, 10, 11)
        End If
    Next i
    
    ' 3. АВТОПОДБОР ШИРИНЫ СТОЛБЦОВ
    wsGraph.Columns.AutoFit
    
    ' 4. РАСЧЕТ ИТОГОВ
    Call CalculateTotals(wsData)
    
    ' Включаем обратно
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.DisplayAlerts = True
    
    MsgBox "График отпусков успешно обновлен!", vbInformation, "Готово"
    Exit Sub
    
ErrorHandler:
    ' Включаем обратно даже при ошибке
    Application.ScreenUpdating = True
    Application.Calculation = xlCalculationAutomatic
    Application.DisplayAlerts = True
    
    MsgBox "Ошибка: " & Err.Description, vbCritical, "Ошибка макроса"
End Sub

' ОЧИСТКА СТАРОГО ГРАФИКА
Private Sub ClearOldSchedule(ws As Worksheet)
    Dim lastCol As Long
    Dim lastRow As Long
    Dim i As Long, j As Long
    
    ' Находим последний столбец
    lastCol = ws.Cells(1, ws.Columns.Count).End(xlToLeft).Column
    
    ' Если есть столбцы кроме A и B, очищаем их
    If lastCol > 2 Then
        For i = 2 To ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
            For j = 3 To lastCol
                With ws.Cells(i, j)
                    .ClearContents
                    .Interior.ColorIndex = xlNone
                    .Font.ColorIndex = xlAutomatic
                    .Font.Bold = False
                End With
            Next j
        Next i
    End If
End Sub

' ОБРАБОТКА ОДНОГО ОТПУСКА
Private Sub ProcessVacation(wsData As Worksheet, wsGraph As Worksheet, _
                           rowNum As Long, startCol As Long, _
                           endCol As Long, daysCol As Long)
    Dim startDate As Date
    Dim endDate As Date
    Dim currentDate As Date
    Dim colNum As Long
    Dim daysCount As Long
    
    ' Пробуем получить даты
    On Error Resume Next
    startDate = CDate(wsData.Cells(rowNum, startCol).Value)
    endDate = CDate(wsData.Cells(rowNum, endCol).Value)
    On Error GoTo 0
    
    ' Если даты валидны
    If startDate > 0 And endDate > 0 And endDate >= startDate Then
        ' Рассчитываем количество дней
        daysCount = DateDiff("d", startDate, endDate) + 1
        wsData.Cells(rowNum, daysCol).Value = daysCount
        
        ' Закрашиваем дни в графике
        currentDate = startDate
        Do While currentDate <= endDate
            colNum = FindDateColumn(wsGraph, currentDate)
            
            If colNum > 0 Then
                With wsGraph.Cells(rowNum, colNum)
                    .Value = "О"
                    .Interior.Color = RGB(144, 238, 144) ' Светло-зеленый
                    .Font.Bold = True
                    .Font.Color = RGB(0, 100, 0) ' Темно-зеленый
                    .HorizontalAlignment = xlCenter
                    .VerticalAlignment = xlCenter
                End With
            End If
            
            currentDate = DateAdd("d", 1, currentDate)
        Loop
    Else
        ' Очищаем поле с днями, если даты не валидны
        wsData.Cells(rowNum, daysCol).ClearContents
    End If
End Sub

' ПОИСК СТОЛБЦА С ДАТОЙ В ГРАФИКЕ
Private Function FindDateColumn(ws As Worksheet, searchDate As Date) As Long
    Dim col As Long
    Dim cellValue As Variant
    
    For col = 3 To ws.Columns.Count
        cellValue = ws.Cells(2, col).Value
        
        ' Проверяем, что в ячейке число (день месяца)
        If IsNumeric(cellValue) Then
            ' Восстанавливаем полную дату
            Dim cellDate As Date
            Dim firstDate As Date
            
            ' Первая дата в графике - 01.01.2026 в столбце C
            firstDate = DateSerial(2026, 1, 1)
            cellDate = DateAdd("d", col - 3, firstDate)
            
            ' Сравниваем с искомой датой
            If Year(cellDate) = Year(searchDate) And _
               Month(cellDate) = Month(searchDate) And _
               Day(cellDate) = Day(searchDate) Then
                FindDateColumn = col
                Exit Function
            End If
        End If
    Next col
    
    FindDateColumn = 0 ' Дата не найдена
End Function

' РАСЧЕТ ИТОГОВЫХ ДНЕЙ ОТПУСКА
Private Sub CalculateTotals(ws As Worksheet)
    Dim lastRow As Long
    Dim totalDays As Long
    Dim i As Long
    
    lastRow = ws.Cells(ws.Rows.Count, 1).End(xlUp).Row
    
    ' Суммируем все дни отпусков
    totalDays = 0
    
    For i = 2 To lastRow
        If IsNumeric(ws.Cells(i, 5).Value) Then totalDays = totalDays + ws.Cells(i, 5).Value
        If IsNumeric(ws.Cells(i, 8).Value) Then totalDays = totalDays + ws.Cells(i, 8).Value
        If IsNumeric(ws.Cells(i, 11).Value) Then totalDays = totalDays + ws.Cells(i, 11).Value
    Next i
    
    ' Выводим итог
    ws.Cells(lastRow + 1, 1).Value = "ИТОГО дней отпуска:"
    ws.Cells(lastRow + 1, 1).Font.Bold = True
    
    ws.Cells(lastRow + 1, 5).Value = totalDays
    ws.Cells(lastRow + 1, 5).Font.Bold = True
    ws.Cells(lastRow + 1, 5).HorizontalAlignment = xlRight
End Sub

' ПРОСТОЙ ТЕСТОВЫЙ МАКРОС
Public Sub TestMacro()
    MsgBox "Макрос работает! Теперь запустите UpdateSchedule.", vbInformation, "Тест"
End Sub
'''
    
    # Сохраняем макрос в отдельный файл
    with open("vacation_macro.txt", "w", encoding="utf-8") as f:
        f.write(macro_code)
    
    print(f"📄 Создан файл с макросом: vacation_macro.txt")
    
    print("\n" + "=" * 70)
    print("📋 ИНСТРУКЦИЯ ПО ИСПОЛЬЗОВАНИЮ:")
    print("=" * 70)
    
    print("\n1. ЗАПУСТИТЕ СКРИПТ:")
    print("   python vacation_creator.py")
    print("   → Создаст файл 'отпуск_график_готовый.xlsx'")
    
    print("\n2. ОТКРОЙТЕ EXCEL ФАЙЛ:")
    print("   - Файл 'отпуск_график_готовый.xlsx'")
    print("   - На листе 'ИНСТРУКЦИЯ' подробные указания")
    
    print("\n3. ДОБАВЬТЕ МАКРОС В EXCEL:")
    print("   - Откройте файл в Excel")
    print("   - Нажмите Alt+F11 (редактор VBA)")
    print("   - Insert → Module")
    print("   - Скопируйте код из 'vacation_macro.txt'")
    print("   - Закройте редактор VBA")
    
    print("\n4. ЗАПУСТИТЕ МАКРОС:")
    print("   - Вернитесь в Excel")
    print("   - Alt+F8 → выберите 'UpdateSchedule' → Выполнить")
    
    print("\n5. РАБОТАЙТЕ С ГРАФИКОМ:")
    print("   - Вносите даты на листе 'ДАННЫЕ'")
    print("   - Запускайте макрос после изменений")
    print("   - График обновится на листе 'ГРАФИК'")

def main():
    try:
        create_vacation_file()
        
        print("\n" + "=" * 70)
        print("✅ ВСЕ ФАЙЛЫ СОЗДАНЫ!")
        print("=" * 70)
        
        print("\n📁 СОЗДАННЫЕ ФАЙЛЫ:")
        print("   1. отпуск_график_готовый.xlsx - основной Excel файл")
        print("   2. vacation_macro.txt - код макроса для копирования")
        
        input("\nНажмите Enter для завершения...")
        
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    main()