from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

def create_vacation_schedule_2026():
    """
    Создает график отпусков на 2026 год с двумя листами:
    1. Основной календарь с отпусками
    2. Лист с данными сотрудников и периодами отпусков
    """
    
    # Создаем все дни 2026 года
    start_date = datetime(2026, 1, 1)
    end_date = datetime(2026, 12, 31)
    
    all_dates = []
    current_date = start_date
    while current_date <= end_date:
        all_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # Праздники 2026 (ст. 112 ТК РФ)
    holidays = [
        # Новогодние каникулы
        *[datetime(2026, 1, d) for d in range(1, 9)],
        datetime(2026, 1, 7),  # Рождество
        datetime(2026, 2, 23),  # 23 февраля
        datetime(2026, 3, 8),   # 8 марта
        datetime(2026, 5, 1),   # 1 мая
        datetime(2026, 5, 9),   # 9 мая
        datetime(2026, 6, 12),  # 12 июня
        datetime(2026, 11, 4),  # 4 ноября
    ]
    
    # Предпраздничные дни (сокращенный рабочий день)
    pre_holidays = [
        datetime(2026, 2, 20),  # перед 23 февраля
        datetime(2026, 3, 7),   # перед 8 марта
        datetime(2026, 5, 8),   # перед 9 мая
        datetime(2026, 6, 11),  # перед 12 июня
        datetime(2026, 11, 3),  # перед 4 ноября
        datetime(2026, 12, 31), # перед Новым годом
    ]
    
    # Создаем новую книгу Excel
    wb = Workbook()
    
    # Удаляем стандартный лист
    wb.remove(wb.active)
    
    # Создаем лист для календаря отпусков
    ws_calendar = wb.create_sheet(title="График отпусков 2026")
    
    # Создаем лист для данных сотрудников
    ws_employees = wb.create_sheet(title="Сотрудники")
    
    # Стили для форматирования
    # Рабочие дни
    workday_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # белый
    workday_font = Font(color="000000", bold=False)  # черный
    
    # Выходные
    weekend_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # светло-серый
    weekend_font = Font(color="000000", bold=False)  # черный
    
    # Праздники
    holiday_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # светло-красный
    holiday_font = Font(color="000000", bold=True)  # черный жирный
    
    # Предпраздничные
    preholiday_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # светло-желтый
    preholiday_font = Font(color="000000", italic=True)  # черный курсив
    
    # Заголовки столбцов
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # синий
    header_font = Font(color="FFFFFF", bold=True)  # белый жирный
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== СОЗДАЕМ ЗАГОЛОВОК КАЛЕНДАРЯ ==========
    
    # Первая строка: Месяцы и дни недели
    ws_calendar['A1'] = "№"
    ws_calendar['A1'].fill = header_fill
    ws_calendar['A1'].font = header_font
    ws_calendar['A1'].alignment = header_alignment
    ws_calendar['A1'].border = thin_border
    
    ws_calendar['B1'] = "ФИО сотрудника"
    ws_calendar['B1'].fill = header_fill
    ws_calendar['B1'].font = header_font
    ws_calendar['B1'].alignment = header_alignment
    ws_calendar['B1'].border = thin_border
    
    ws_calendar['C1'] = "Должность"
    ws_calendar['C1'].fill = header_fill
    ws_calendar['C1'].font = header_font
    ws_calendar['C1'].alignment = header_alignment
    ws_calendar['C1'].border = thin_border
    
    ws_calendar['D1'] = "Отдел"
    ws_calendar['D1'].fill = header_fill
    ws_calendar['D1'].font = header_font
    ws_calendar['D1'].alignment = header_alignment
    ws_calendar['D1'].border = thin_border
    
    # Заполняем календарь по столбцам, начиная с E1
    col_index = 5  # Начинаем с колонки E
    
    # Словарь для названий месяцев
    months_ru = {
        1: "Янв", 2: "Фев", 3: "Мар", 4: "Апр", 5: "Май", 6: "Июн",
        7: "Июл", 8: "Авг", 9: "Сен", 10: "Окт", 11: "Ноя", 12: "Дек"
    }
    
    days_short = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    current_month = 0
    month_cols = {}  # Для группировки колонок по месяцам
    
    for i, date in enumerate(all_dates, 1):
        col_letter = get_column_letter(col_index)
        
        # Определяем тип дня
        weekday = date.weekday()  # 0-пн, 6-вс
        is_weekend = weekday >= 5
        is_holiday = date in holidays
        is_preholiday = date in pre_holidays
        
        # Форматируем ячейку
        cell = ws_calendar[f"{col_letter}1"]
        
        if is_holiday:
            cell.value = f"{date.day}\n✶"
            cell.fill = holiday_fill
            cell.font = holiday_font
        elif is_preholiday:
            cell.value = f"{date.day}\n◐"
            cell.fill = preholiday_fill
            cell.font = preholiday_font
        elif is_weekend:
            cell.value = f"{date.day}\n{days_short[weekday]}"
            cell.fill = weekend_fill
            cell.font = weekend_font
        else:
            cell.value = f"{date.day}\n{days_short[weekday]}"
            cell.fill = workday_fill
            cell.font = workday_font
        
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
        # Вторая строка: номер месяца
        cell2 = ws_calendar[f"{col_letter}2"]
        if date.month != current_month:
            cell2.value = months_ru[date.month]
            cell2.fill = header_fill
            cell2.font = header_font
            cell2.alignment = header_alignment
            cell2.border = thin_border
            current_month = date.month
            # Запоминаем начало месяца для объединения
            if date.month not in month_cols:
                month_cols[date.month] = col_index
        else:
            cell2.value = ""
        
        # Настраиваем ширину столбца
        ws_calendar.column_dimensions[col_letter].width = 4
        
        col_index += 1
    
    # Объединяем ячейки с названиями месяцев
    for month, start_col in month_cols.items():
        # Находим последний день месяца
        if month == 12:
            end_date_month = datetime(2026, 12, 31)
        else:
            end_date_month = datetime(2026, month + 1, 1) - timedelta(days=1)
        
        # Находим индекс последнего дня месяца
        end_col = start_col
        for col in range(start_col, col_index):
            col_letter = get_column_letter(col)
            # Проверяем, что это день нужного месяца
            date_index = col - 5
            if date_index < len(all_dates):
                date = all_dates[date_index]
                if date.month == month:
                    end_col = col
        
        if end_col >= start_col:
            start_letter = get_column_letter(start_col)
            end_letter = get_column_letter(end_col)
            if start_letter != end_letter:
                ws_calendar.merge_cells(f"{start_letter}2:{end_letter}2")
    
    # Третья строка: можно добавить рабочее время или другую информацию
    ws_calendar['A3'] = ""
    ws_calendar['B3'] = ""
    ws_calendar['C3'] = ""
    ws_calendar['D3'] = ""
    
    # Заполняем третью строку для календарных дней
    for col in range(5, col_index):
        col_letter = get_column_letter(col)
        cell = ws_calendar[f"{col_letter}3"]
        cell.value = ""  # Можно добавить "8ч" или другую информацию
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Устанавливаем ширину для первых столбцов
    ws_calendar.column_dimensions['A'].width = 4      # №
    ws_calendar.column_dimensions['B'].width = 25     # ФИО
    ws_calendar.column_dimensions['C'].width = 20     # Должность
    ws_calendar.column_dimensions['D'].width = 15     # Отдел
    
    # ========== СОЗДАЕМ ЛИСТ С ДАННЫМИ СОТРУДНИКОВ ==========
    
    # Заголовки для листа сотрудников
    ws_employees['A1'] = "№"
    ws_employees['A1'].fill = header_fill
    ws_employees['A1'].font = header_font
    ws_employees['A1'].alignment = header_alignment
    ws_employees['A1'].border = thin_border
    
    ws_employees['B1'] = "ФИО сотрудника"
    ws_employees['B1'].fill = header_fill
    ws_employees['B1'].font = header_font
    ws_employees['B1'].alignment = header_alignment
    ws_employees['B1'].border = thin_border
    
    ws_employees['C1'] = "Должность"
    ws_employees['C1'].fill = header_fill
    ws_employees['C1'].font = header_font
    ws_employees['C1'].alignment = header_alignment
    ws_employees['C1'].border = thin_border
    
    ws_employees['D1'] = "Отдел"
    ws_employees['D1'].fill = header_fill
    ws_employees['D1'].font = header_font
    ws_employees['D1'].alignment = header_alignment
    ws_employees['D1'].border = thin_border
    
    # Заголовки для периодов отпусков
    ws_employees['E1'] = "Периоды отпусков"
    ws_employees['E1'].fill = header_fill
    ws_employees['E1'].font = header_font
    ws_employees['E1'].alignment = header_alignment
    ws_employees['E1'].border = thin_border
    
    # Объединяем ячейку для заголовка периодов
    ws_employees.merge_cells('E1:F1')
    
    # Подзаголовки для периодов
    ws_employees['E2'] = "Начало отпуска"
    ws_employees['E2'].fill = header_fill
    ws_employees['E2'].font = header_font
    ws_employees['E2'].alignment = header_alignment
    ws_employees['E2'].border = thin_border
    
    ws_employees['F2'] = "Конец отпуска"
    ws_employees['F2'].fill = header_fill
    ws_employees['F2'].font = header_font
    ws_employees['F2'].alignment = header_alignment
    ws_employees['F2'].border = thin_border
    
    # Добавляем примерные данные сотрудников
    employees_data = [
        ["Иванов Иван Иванович", "Менеджер", "Отдел продаж"],
        ["Петров Петр Петрович", "Разработчик", "IT отдел"],
        ["Сидорова Мария Владимировна", "Бухгалтер", "Бухгалтерия"],
        ["Козлов Алексей Николаевич", "HR-менеджер", "Отдел кадров"],
        ["Морозова Елена Сергеевна", "Дизайнер", "Отдел маркетинга"],
    ]
    
    # Заполняем данными сотрудников
    for i, employee in enumerate(employees_data, 1):
        # Номер
        ws_employees[f'A{i+2}'] = i
        ws_employees[f'A{i+2}'].border = thin_border
        ws_employees[f'A{i+2}'].alignment = Alignment(horizontal="center", vertical="center")
        
        # ФИО
        ws_employees[f'B{i+2}'] = employee[0]
        ws_employees[f'B{i+2}'].border = thin_border
        ws_employees[f'B{i+2}'].alignment = Alignment(horizontal="left", vertical="center")
        
        # Должность
        ws_employees[f'C{i+2}'] = employee[1]
        ws_employees[f'C{i+2}'].border = thin_border
        ws_employees[f'C{i+2}'].alignment = Alignment(horizontal="left", vertical="center")
        
        # Отдел
        ws_employees[f'D{i+2}'] = employee[2]
        ws_employees[f'D{i+2}'].border = thin_border
        ws_employees[f'D{i+2}'].alignment = Alignment(horizontal="left", vertical="center")
        
        # Периоды отпусков (пустые ячейки для заполнения)
        ws_employees[f'E{i+2}'] = ""
        ws_employees[f'E{i+2}'].border = thin_border
        ws_employees[f'E{i+2}'].alignment = Alignment(horizontal="center", vertical="center")
        
        ws_employees[f'F{i+2}'] = ""
        ws_employees[f'F{i+2}'].border = thin_border
        ws_employees[f'F{i+2}'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Устанавливаем ширину столбцов
    ws_employees.column_dimensions['A'].width = 5
    ws_employees.column_dimensions['B'].width = 30
    ws_employees.column_dimensions['C'].width = 20
    ws_employees.column_dimensions['D'].width = 20
    ws_employees.column_dimensions['E'].width = 15
    ws_employees.column_dimensions['F'].width = 15
    
    # ========== ДОБАВЛЯЕМ СЛУЖЕБНУЮ ИНФОРМАЦИЮ ==========
    
    # Добавляем лист с легендой
    ws_legend = wb.create_sheet(title="Легенда")
    
    legend_data = [
        ["Обозначения в календаре:", ""],
        ["", ""],
        ["Цвет", "Тип дня", "Обозначение"],
        ["Белый", "Рабочий день", "Число + Пн/Вт/Ср/Чт/Пт"],
        ["Серый", "Выходной", "Число + Сб/Вс"],
        ["Красный", "Праздничный", "Число + ✶"],
        ["Желтый", "Предпраздничный", "Число + ◐"],
        ["", ""],
        ["Статистика 2026:", ""],
        ["Всего дней", len(all_dates)],
        ["Рабочих дней", len([d for d in all_dates if d.weekday() < 5 and d not in holidays])],
        ["Выходных", len([d for d in all_dates if d.weekday() >= 5])],
        ["Праздничных", len(holidays)],
    ]
    
    for i, row in enumerate(legend_data, 1):
        for j, value in enumerate(row, 1):
            cell = ws_legend.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i <= 3 or i == 8 or i == 9:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    
    # Добавляем лист с праздниками
    ws_holidays = wb.create_sheet(title="Праздники")
    
    holidays_list = [
        ["Дата", "Праздник", "Тип дня"],
        ["01.01.2026", "Новый год", "Праздничный"],
        ["02.01.2026", "Новогодние каникулы", "Праздничный"],
        ["03.01.2026", "Новогодние каникулы", "Праздничный"],
        ["04.01.2026", "Новогодние каникулы", "Праздничный"],
        ["05.01.2026", "Новогодние каникулы", "Праздничный"],
        ["06.01.2026", "Новогодние каникулы", "Праздничный"],
        ["07.01.2026", "Рождество Христово", "Праздничный"],
        ["08.01.2026", "Новогодние каникулы", "Праздничный"],
        ["23.02.2026", "День защитника Отечества", "Праздничный"],
        ["08.03.2026", "Международный женский день", "Праздничный"],
        ["01.05.2026", "Праздник Весны и Труда", "Праздничный"],
        ["09.05.2026", "День Победы", "Праздничный"],
        ["12.06.2026", "День России", "Праздничный"],
        ["04.11.2026", "День народного единства", "Праздничный"],
    ]
    
    for i, row in enumerate(holidays_list, 1):
        for j, value in enumerate(row, 1):
            cell = ws_holidays.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
    
    # Сохраняем файл
    output_file = "график_отпусков_2026.xlsx"
    wb.save(output_file)
    
    print(f"✅ Файл создан: {output_file}")
    print(f"📊 Структура файла:")
    print(f"   1. Лист 'График отпусков 2026' - основной с календарем в шапке")
    print(f"   2. Лист 'Сотрудники' - данные сотрудников и периоды отпусков")
    print(f"   3. Лист 'Легенда' - расшифровка обозначений")
    print(f"   4. Лист 'Праздники' - список праздничных дней")
    print(f"\n📅 Календарь оформлен как заголовки столбцов:")
    print(f"   - Строка 1: число дня и обозначение (Пн/Вс/✶/◐)")
    print(f"   - Строка 2: месяц (объединен по всем дням месяца)")
    print(f"   - Строка 3: зарезервирована для дополнительной информации")
    print(f"\n📋 На листе 'Сотрудники' можно вводить данные:")
    print(f"   - В столбце B - ФИО сотрудников")
    print(f"   - В столбце E - даты начала отпусков")
    print(f"   - В столбце F - даты окончания отпусков")
    
    return output_file

# Запускаем создание графика отпусков
if __name__ == "__main__":
    filename = create_vacation_schedule_2026()
    
    print(f"\n🚀 Инструкция по использованию:")
    print(f"1. Файл '{filename}' готов к использованию")
    print(f"2. На листе 'Сотрудники' заполните данные сотрудников")
    print(f"3. Введите даты начала и окончания отпусков в соответствующие столбцы")
    print(f"4. Для автоматического заполнения календаря используйте функцию обновления")