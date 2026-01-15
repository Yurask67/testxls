#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ГЕНЕРАТОР ГРАФИКА ОТПУСКОВ 2026
С ДИНАМИЧЕСКИМИ ФОРМУЛАМИ ДЛЯ АВТОМАТИЧЕСКОГО ОБНОВЛЕНИЯ
"""

import os
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

def get_russian_calendar_2026():
    """Возвращает производственный календарь России на 2026 год"""
    
    # Праздничные дни (нерабочие)
    holidays = [
        # Новогодние каникулы и Рождество
        (2026, 1, 1), (2026, 1, 2), (2026, 1, 3), (2026, 1, 4),
        (2026, 1, 5), (2026, 1, 6), (2026, 1, 7), (2026, 1, 8),
        (2026, 1, 9),  # 9 января - дополнительный выходной
        
        # 23 Февраля
        (2026, 2, 23),
        
        # 8 Марта
        (2026, 3, 8),
        
        # 1 Мая
        (2026, 5, 1),
        
        # 9 Мая
        (2026, 5, 9),
        
        # 12 Июня
        (2026, 6, 12),
        
        # 4 Ноября
        (2026, 11, 4),
    ]
    
    # Предпраздничные дни (сокращенные на 1 час)
    pre_holidays = [
        (2026, 2, 20),  # Пятница перед 23 февраля
        (2026, 3, 7),   # Суббота перед 8 марта (рабочая)
        (2026, 5, 8),   # Пятница перед 9 мая
        (2026, 6, 11),  # Пятница перед 12 июня
        (2026, 11, 3),  # Вторник перед 4 ноября
        (2026, 12, 31), # Четверг перед Новым годом
    ]
    
    # Рабочие субботы (переносы)
    working_saturdays = [
        (2026, 2, 21),  # Суббота (рабочая вместо понедельника)
        (2026, 11, 14), # Суббота (рабочая вместо понедельника)
    ]
    
    # Создаем календарь на весь год
    calendar = {}
    start_date = datetime(2026, 1, 1)
    
    for i in range(365 + 1):  # +1 для високосного 2026
        current_date = start_date + timedelta(days=i)
        if current_date.year > 2026:
            break
            
        date_key = current_date.date()
        weekday = current_date.weekday()  # 0=пн, 6=вс
        
        # Определяем тип дня
        is_holiday = (current_date.year, current_date.month, current_date.day) in holidays
        is_pre_holiday = (current_date.year, current_date.month, current_date.day) in pre_holidays
        is_working_saturday = (current_date.year, current_date.month, current_date.day) in working_saturdays
        
        if is_holiday:
            day_type = "holiday"
            day_name = "Праздник"
        elif is_pre_holiday:
            day_type = "pre_holiday"
            day_name = "Предпр"
        elif is_working_saturday:
            day_type = "work_saturday"
            day_name = "Раб.сб"
        elif weekday >= 5:  # Суббота или воскресенье
            day_type = "weekend"
            day_name = "Выходной"
        else:
            day_type = "workday"
            day_name = "Рабочий"
        
        # Название праздника
        holiday_name = ""
        if is_holiday:
            if current_date.month == 1 and current_date.day <= 9:
                holiday_name = "Новогодние каникулы"
            elif current_date.month == 1 and current_date.day == 7:
                holiday_name = "Рождество"
            elif current_date.month == 2 and current_date.day == 23:
                holiday_name = "День защитника Отечества"
            elif current_date.month == 3 and current_date.day == 8:
                holiday_name = "Международный женский день"
            elif current_date.month == 5 and current_date.day == 1:
                holiday_name = "Праздник Весны и Труда"
            elif current_date.month == 5 and current_date.day == 9:
                holiday_name = "День Победы"
            elif current_date.month == 6 and current_date.day == 12:
                holiday_name = "День России"
            elif current_date.month == 11 and current_date.day == 4:
                holiday_name = "День народного единства"
        
        calendar[date_key] = {
            'date': current_date,
            'day': current_date.day,
            'month': current_date.month,
            'year': current_date.year,
            'weekday': weekday,
            'day_type': day_type,
            'day_name': day_name,
            'holiday_name': holiday_name,
            'is_working': day_type in ['workday', 'work_saturday', 'pre_holiday']
        }
    
    return calendar

def create_dynamic_vacation_schedule():
    """Создает график отпусков с ДИНАМИЧЕСКИМИ ФОРМУЛАМИ для автообновления"""
    
    print("=" * 70)
    print("ГЕНЕРАТОР ГРАФИКА ОТПУСКОВ 2026")
    print("С ДИНАМИЧЕСКИМИ ФОРМУЛАМИ ДЛЯ АВТООБНОВЛЕНИЯ")
    print("=" * 70)
    
    # 1. ГЕНЕРИРУЕМ КАЛЕНДАРЬ
    print("\n📅 Генерирую производственный календарь РФ на 2026 год...")
    calendar = get_russian_calendar_2026()
    
    # 2. ИМЯ ФАЙЛА
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"отпуск_динамика_2026_{timestamp}.xlsx"
    
    print(f"\n📁 Имя файла: {default_name}")
    user_input = input("Введите свое имя файла (или Enter для умолчания): ").strip()
    
    if user_input:
        if not user_input.endswith('.xlsx'):
            user_input += '.xlsx'
        filename = user_input
    else:
        filename = default_name
    
    # Проверка существования файла
    if os.path.exists(filename):
        print(f"⚠️ Файл '{filename}' уже существует!")
        overwrite = input("Перезаписать? (y/n): ").lower()
        if overwrite != 'y':
            print("❌ Отменено")
            return
    
    # 3. СОЗДАЕМ КНИГУ EXCEL
    print("\n🔄 Создаю файл Excel с динамическими формулами...")
    wb = Workbook()
    
    # Удаляем дефолтный лист
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # 4. СОЗДАЕМ ЛИСТ СОТРУДНИКОВ
    print("👥 Создаю лист сотрудников...")
    ws_employees = wb.create_sheet(title="СОТРУДНИКИ")
    
    # Стили
    header_fill = PatternFill(start_color="1F497D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки листа сотрудников
    headers = [
        "№", "ФАМИЛИЯ ИМЯ ОТЧЕСТВО",
        "ОТПУСК 1", "ОТПУСК 1", "ОТПУСК 1",
        "ОТПУСК 2", "ОТПУСК 2", "ОТПУСК 2",
        "ОТПУСК 3", "ОТПУСК 3", "ОТПУСК 3"
    ]
    
    for col, header in enumerate(headers, 1):
        ws_employees.cell(row=1, column=col, value=header)
    
    # Объединяем ячейки для заголовков отпусков
    ws_employees.merge_cells('C1:E1')
    ws_employees.merge_cells('F1:H1')
    ws_employees.merge_cells('I1:K1')
    
    # Заголовки второго ряда
    sub_headers = ["", "",
                  "Начало", "Конец", "Дней",
                  "Начало", "Конец", "Дней",
                  "Начало", "Конец", "Дней"]
    
    for col, header in enumerate(sub_headers, 1):
        if header:
            ws_employees.cell(row=2, column=col, value=header)
    
    # Применяем стили к заголовкам
    for row in [1, 2]:
        for col in range(1, 12):
            cell = ws_employees.cell(row=row, column=col)
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
    
    # Настраиваем ширину столбцов
    column_widths = [5, 30, 12, 12, 8, 12, 12, 8, 12, 12, 8]
    for i, width in enumerate(column_widths, 1):
        ws_employees.column_dimensions[get_column_letter(i)].width = width
    
    # ДАННЫЕ СОТРУДНИКОВ (пример с тестовыми отпусками)
    employees_data = [
        {
            "name": "ИВАНОВ ИВАН ИВАНОВИЧ",
            "vacations": [
                {"start": "2026-01-10", "end": "2026-01-25"},  # 16 дней
                {"start": "2026-07-15", "end": "2026-08-01"},  # 18 дней
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "ПЕТРОВ ПЕТР ПЕТРОВИЧ",
            "vacations": [
                {"start": "2026-02-15", "end": "2026-02-25"},  # 11 дней
                {"start": "2026-09-01", "end": "2026-09-14"},  # 14 дней
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "СИДОРОВА МАРИЯ ВЛАДИМИРОВНА",
            "vacations": [
                {"start": "2026-03-01", "end": "2026-03-14"},  # 14 дней
                {"start": "2026-10-10", "end": "2026-10-20"},  # 11 дней
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "КОЗЛОВ АЛЕКСЕЙ НИКОЛАЕВИЧ",
            "vacations": [
                {"start": "2026-04-01", "end": "2026-04-10"},  # 10 дней
                {"start": "2026-11-01", "end": "2026-11-10"},  # 10 дней
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "МОРОЗОВА ЕЛЕНА СЕРГЕЕВНА",
            "vacations": [
                {"start": "2026-05-10", "end": "2026-05-24"},  # 15 дней
                {"start": "2026-12-15", "end": "2026-12-31"},  # 17 дней
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "НИКОЛАЕВ АНДРЕЙ ВИКТОРОВИЧ",
            "vacations": [
                {"start": "2026-06-01", "end": "2026-06-14"},  # 14 дней
                {"start": "", "end": ""},
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "ОРЛОВА ОЛЬГА ИГОРЕВНА",
            "vacations": [
                {"start": "2026-07-01", "end": "2026-07-10"},  # 10 дней
                {"start": "", "end": ""},
                {"start": "", "end": ""}
            ]
        },
        {
            "name": "ВОЛКОВ ДМИТРИЙ АЛЕКСАНДРОВИЧ",
            "vacations": [
                {"start": "2026-08-15", "end": "2026-08-31"},  # 17 дней
                {"start": "", "end": ""},
                {"start": "", "end": ""}
            ]
        }
    ]
    
    # Заполняем данные сотрудников
    for i, emp in enumerate(employees_data, start=3):
        # Номер
        ws_employees.cell(row=i, column=1, value=i-2)
        ws_employees.cell(row=i, column=1).alignment = center_align
        ws_employees.cell(row=i, column=1).border = thin_border
        
        # ФИО
        ws_employees.cell(row=i, column=2, value=emp["name"])
        ws_employees.cell(row=i, column=2).alignment = Alignment(vertical="center")
        ws_employees.cell(row=i, column=2).border = thin_border
        
        # Даты отпусков
        vacation_cols = [(3, 4), (6, 7), (9, 10)]  # Пары столбцов для дат
        
        for j, (start_col, end_col) in enumerate(vacation_cols):
            if j < len(emp["vacations"]):
                vac = emp["vacations"][j]
                
                # Дата начала
                if vac["start"]:
                    try:
                        ws_employees.cell(row=i, column=start_col, 
                                        value=datetime.strptime(vac["start"], "%Y-%m-%d"))
                    except ValueError:
                        ws_employees.cell(row=i, column=start_col, value=vac["start"])
                
                # Дата окончания
                if vac["end"]:
                    try:
                        ws_employees.cell(row=i, column=end_col, 
                                        value=datetime.strptime(vac["end"], "%Y-%m-%d"))
                    except ValueError:
                        ws_employees.cell(row=i, column=end_col, value=vac["end"])
        
        # Формулы для расчета дней отпуска
        formula_cols = [5, 8, 11]  # Столбцы для формул
        
        for j, formula_col in enumerate(formula_cols):
            start_col = formula_col - 2  # C, F, I
            end_col = formula_col - 1    # D, G, J
            
            formula = f'=IF(AND({get_column_letter(start_col)}{i}<>"",{get_column_letter(end_col)}{i}<>""),{get_column_letter(end_col)}{i}-{get_column_letter(start_col)}{i}+1,"")'
            ws_employees.cell(row=i, column=formula_col, value=formula)
            ws_employees.cell(row=i, column=formula_col).number_format = '0'
        
        # Применяем границы и форматирование
        for col in range(3, 12):  # Столбцы C-K
            cell = ws_employees.cell(row=i, column=col)
            cell.border = thin_border
            cell.alignment = center_align
            
            # Форматирование дат
            if col in [3, 4, 6, 7, 9, 10]:  # Столбцы с датами
                cell.number_format = 'DD.MM.YYYY'
        
        # Закрашиваем строку через одну
        if i % 2 == 0:
            row_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
            for col in range(1, 12):
                ws_employees.cell(row=i, column=col).fill = row_fill
    
    # Формула для итогового количества дней отпуска
    last_row = len(employees_data) + 2
    ws_employees.cell(row=last_row+1, column=1, value="ИТОГО дней отпуска:")
    ws_employees.cell(row=last_row+1, column=1).font = Font(bold=True)
    
    formula_total = f'=SUM(E3:E{last_row},H3:H{last_row},K3:K{last_row})'
    ws_employees.cell(row=last_row+1, column=5, value=formula_total)
    ws_employees.cell(row=last_row+1, column=5).font = Font(bold=True)
    ws_employees.cell(row=last_row+1, column=5).number_format = '0'
    
    # 5. СОЗДАЕМ ЛИСТ С ГРАФИКОМ ОТПУСКОВ (С ДИНАМИЧЕСКИМИ ФОРМУЛАМИ)
    print("📅 Создаю лист с графиком отпусков (динамические формулы)...")
    ws_schedule = wb.create_sheet(title="ГРАФИК ОТПУСКОВ")
    
    # Стили для графика
    month_fills = {
        1: PatternFill(start_color="4F81BD", fill_type="solid"),
        2: PatternFill(start_color="8064A2", fill_type="solid"),
        3: PatternFill(start_color="9BBB59", fill_type="solid"),
        4: PatternFill(start_color="C0504D", fill_type="solid"),
        5: PatternFill(start_color="F79646", fill_type="solid"),
        6: PatternFill(start_color="1F497D", fill_type="solid"),
        7: PatternFill(start_color="948A54", fill_type="solid"),
        8: PatternFill(start_color="31869B", fill_type="solid"),
        9: PatternFill(start_color="E26B0A", fill_type="solid"),
        10: PatternFill(start_color="60497A", fill_type="solid"),
        11: PatternFill(start_color="C00000", fill_type="solid"),
        12: PatternFill(start_color="366092", fill_type="solid"),
    }
    
    day_type_fills = {
        'workday': PatternFill(start_color="FFFFFF", fill_type="solid"),
        'weekend': PatternFill(start_color="E6E6E6", fill_type="solid"),
        'holiday': PatternFill(start_color="FF9999", fill_type="solid"),
        'pre_holiday': PatternFill(start_color="FFFF99", fill_type="solid"),
        'work_saturday': PatternFill(start_color="CCFFCC", fill_type="solid")
    }
    
    day_type_fonts = {
        'workday': Font(color="000000", size=9),
        'weekend': Font(color="000000", size=9),
        'holiday': Font(color="000000", bold=True, size=9),
        'pre_holiday': Font(color="000000", italic=True, size=9),
        'work_saturday': Font(color="006600", bold=True, size=9)
    }
    
    # Заголовки для графика
    ws_schedule['A1'] = "№"
    ws_schedule['A1'].fill = header_fill
    ws_schedule['A1'].font = header_font
    ws_schedule['A1'].alignment = center_align
    ws_schedule['A1'].border = thin_border
    ws_schedule.column_dimensions['A'].width = 5
    
    ws_schedule['B1'] = "ФИО СОТРУДНИКА"
    ws_schedule['B1'].fill = header_fill
    ws_schedule['B1'].font = header_font
    ws_schedule['B1'].alignment = center_align
    ws_schedule['B1'].border = thin_border
    ws_schedule.column_dimensions['B'].width = 30
    
    # Группируем дни по месяцам
    months_data = {}
    for date_obj, day_info in calendar.items():
        month = day_info['month']
        if month not in months_data:
            months_data[month] = []
        months_data[month].append(day_info)
    
    # Сортируем месяцы
    sorted_months = sorted(months_data.keys())
    
    # Создаем маппинг дата -> столбец
    date_column_map = {}
    current_col = 3  # Начинаем с колонки C
    
    # Названия месяцев
    month_names = {
        1: "ЯНВ", 2: "ФЕВ", 3: "МАР", 4: "АПР",
        5: "МАЙ", 6: "ИЮН", 7: "ИЮЛ", 8: "АВГ",
        9: "СЕН", 10: "ОКТ", 11: "НОЯ", 12: "ДЕК"
    }
    
    # Дни недели сокращенные
    weekday_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    
    # Заполняем каждый месяц
    for month_num in sorted_months:
        month_days = months_data[month_num]
        
        # Объединяем ячейки для названия месяца
        start_col = current_col
        end_col = current_col + len(month_days) - 1
        
        start_letter = get_column_letter(start_col)
        end_letter = get_column_letter(end_col)
        
        ws_schedule.merge_cells(f"{start_letter}1:{end_letter}1")
        
        # Название месяца
        month_cell = ws_schedule[f"{start_letter}1"]
        month_cell.value = month_names[month_num]
        month_cell.fill = month_fills[month_num]
        month_cell.font = header_font
        month_cell.alignment = center_align
        month_cell.border = thin_border
        
        # Заполняем дни месяца
        for i, day_info in enumerate(month_days):
            col = current_col + i
            
            # Сохраняем соответствие дата -> столбец
            date_key = day_info['date'].date()
            date_column_map[date_key] = col
            
            # Строка 2: число дня (скрытая дата для формул)
            date_cell = ws_schedule.cell(row=2, column=col)
            date_cell.value = day_info['date']  # Сохраняем полную дату
            date_cell.number_format = 'DD'  # Показываем только день
            date_cell.alignment = center_align
            date_cell.font = Font(bold=True, size=9)
            date_cell.border = thin_border
            date_cell.fill = day_type_fills[day_info['day_type']]
            
            # Строка 3: день недели + обозначение
            weekday = weekday_names[day_info['weekday']]
            
            # Добавляем символы для особых дней
            symbol = ""
            if day_info['day_type'] == 'holiday':
                symbol = " ✶"
            elif day_info['day_type'] == 'pre_holiday':
                symbol = " ◐"
            elif day_info['day_type'] == 'work_saturday':
                symbol = " ⚒"
            
            day_name_cell = ws_schedule.cell(row=3, column=col, value=f"{weekday}{symbol}")
            day_name_cell.alignment = center_align
            day_name_cell.font = day_type_fonts[day_info['day_type']]
            day_name_cell.border = thin_border
            day_name_cell.fill = day_type_fills[day_info['day_type']]
            
            # Настраиваем ширину столбца
            col_letter = get_column_letter(col)
            ws_schedule.column_dimensions[col_letter].width = 4.5
        
        current_col += len(month_days)
    
    # Добавляем сотрудников на лист графика и создаем ДИНАМИЧЕСКИЕ ФОРМУЛЫ
    print("🎯 Создаю динамические формулы для автоматического обновления...")
    
    # Цвет для отпусков (светло-зеленый) - будет через условное форматирование
    vacation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    vacation_font = Font(bold=True, color="006400")  # Темно-зеленый
    
    for i, emp in enumerate(employees_data, start=4):
        # Номер
        ws_schedule.cell(row=i, column=1, value=i-3)
        ws_schedule.cell(row=i, column=1).alignment = center_align
        ws_schedule.cell(row=i, column=1).border = thin_border
        
        # ФИО
        ws_schedule.cell(row=i, column=2, value=emp["name"])
        ws_schedule.cell(row=i, column=2).alignment = Alignment(vertical="center")
        ws_schedule.cell(row=i, column=2).border = thin_border
        
        # Закрашиваем строку через одну (основной фон)
        if i % 2 == 0:
            row_fill = PatternFill(start_color="F8F8F8", fill_type="solid")
            for col in range(1, current_col):
                ws_schedule.cell(row=i, column=col).fill = row_fill
    
    # 6. ДОБАВЛЯЕМ УСЛОВНОЕ ФОРМАТИРОВАНИЕ ДЛЯ ДИНАМИЧЕСКОГО ОТОБРАЖЕНИЯ ОТПУСКОВ
    print("✨ Добавляю условное форматирование...")
    
    # Определяем диапазон для условного форматирования
    last_row_schedule = len(employees_data) + 3
    last_col = current_col - 1
    
    start_col_letter = get_column_letter(3)  # C
    end_col_letter = get_column_letter(last_col)
    
    range_address = f"{start_col_letter}4:{end_col_letter}{last_row_schedule}"
    
    # Создаем правила условного форматирования для каждого сотрудника
    for i, emp in enumerate(employees_data, start=4):
        employee_row = i
        employee_sheet_row = employee_row - 1  # На листе СОТРУДНИКИ
        
        # Создаем формулу для условного форматирования
        # Проверяем все 3 возможных отпуска
        formula_parts = []
        
        for vac_idx in range(3):  # для 3х возможных отпусков
            # Столбцы на листе СОТРУДНИКИ
            start_col_emp = get_column_letter(3 + vac_idx * 3)  # C, F, I
            end_col_emp = get_column_letter(4 + vac_idx * 3)    # D, G, J
            
            # Формула проверки: дата в столбце >= начала отпуска И дата <= окончания отпуска
            # И проверяем, что даты отпуска не пустые
            formula_part = f'AND(СОТРУДНИКИ!${start_col_emp}${employee_sheet_row}<>"",СОТРУДНИКИ!${end_col_emp}${employee_sheet_row}<>"",$C2>=СОТРУДНИКИ!${start_col_emp}${employee_sheet_row},$C2<=СОТРУДНИКИ!${end_col_emp}${employee_sheet_row})'
            formula_parts.append(formula_part)
        
        # Объединяем все проверки через OR
        if formula_parts:
            full_formula = f'=OR({",".join(formula_parts)})'
            
            # Создаем правило условного форматирования для этой строки
            rule = FormulaRule(
                formula=[full_formula],
                fill=vacation_fill,
                font=vacation_font
            )
            
            # Применяем правило к строке сотрудника
            row_range = f"{start_col_letter}{employee_row}:{end_col_letter}{employee_row}"
            ws_schedule.conditional_formatting.add(row_range, rule)
            
            # Также добавляем формулу в каждую ячейку для отображения "О"
            for col in range(3, last_col + 1):
                cell = ws_schedule.cell(row=employee_row, column=col)
                
                # Создаем формулу для отображения "О"
                formula_parts_display = []
                for vac_idx in range(3):
                    start_col_emp = get_column_letter(3 + vac_idx * 3)
                    end_col_emp = get_column_letter(4 + vac_idx * 3)
                    
                    col_letter = get_column_letter(col)
                    formula_part = f'IF(AND(СОТРУДНИКИ!${start_col_emp}${employee_sheet_row}<>"",СОТРУДНИКИ!${end_col_emp}${employee_sheet_row}<>"",${col_letter}$2>=СОТРУДНИКИ!${start_col_emp}${employee_sheet_row},${col_letter}$2<=СОТРУДНИКИ!${end_col_emp}${employee_sheet_row}),"О","")'
                    formula_parts_display.append(formula_part)
                
                if formula_parts_display:
                    display_formula = f'=IF(OR({",".join(formula_parts_display)}),"О","")'
                    cell.value = display_formula
                    cell.alignment = center_align
    
    # 7. ДОБАВЛЯЕМ ФОРМУЛЫ ДЛЯ ДИНАМИЧЕСКОГО ОБНОВЛЕНИЯ
    print("🔧 Настраиваю динамическое обновление...")
    
    # Добавляем примечание о динамическом обновлении
    note_row = len(employees_data) + 5
    ws_schedule.cell(row=note_row, column=1, value="💡 ДИНАМИЧЕСКОЕ ОБНОВЛЕНИЕ:")
    ws_schedule.cell(row=note_row, column=1).font = Font(bold=True, color="1F497D", size=11)
    
    ws_schedule.cell(row=note_row+1, column=1, value="✅ Отпуска автоматически обновляются при изменении дат")
    ws_schedule.cell(row=note_row+2, column=1, value="✅ Не нужно пересоздавать файл или перезагружать")
    ws_schedule.cell(row=note_row+3, column=1, value="✅ Просто нажмите F9 для пересчета формул")
    ws_schedule.cell(row=note_row+4, column=1, value="✅ Или измените любую ячейку - формулы обновятся автоматически")
    
    # 8. СОЗДАЕМ ЛИСТ С ЛЕГЕНДОЙ (обновленная)
    print("📝 Создаю лист с легендой...")
    ws_legend = wb.create_sheet(title="ЛЕГЕНДА")
    
    # Заголовок
    ws_legend['A1'] = "ЛЕГЕНДА - ОБОЗНАЧЕНИЯ В ГРАФИКЕ"
    ws_legend['A1'].font = Font(bold=True, size=14, color="1F497D")
    ws_legend.merge_cells('A1:C1')
    
    # Обозначения
    legend_data = [
        ["Обозначение", "Тип дня", "Описание"],
        ["Белый фон", "Рабочий день", "Обычный рабочий день (пн-пт)"],
        ["Серый фон", "Выходной день", "Суббота, воскресенье"],
        ["Красный фон + ✶", "Праздничный день", "Государственный праздник"],
        ["Желтый фон + ◐", "Предпраздничный", "Сокращенный рабочий день"],
        ["Зеленый фон + ⚒", "Рабочая суббота", "Перенесенная рабочая суббота"],
        ["Светло-зеленый + О", "Отпуск", "Период отпуска сотрудника"],
        ["", "", ""],
        ["📊 ДИНАМИЧЕСКОЕ ОБНОВЛЕНИЕ:", "", ""],
        ["✅ Автоматически обновляется при изменении дат отпусков", "", ""],
        ["✅ Не нужно закрывать/открывать файл", "", ""],
        ["✅ Нажмите F9 для принудительного пересчета", "", ""],
        ["✅ Или измените любую ячейку", "", ""],
        ["", "", ""],
        ["📋 ИНСТРУКЦИЯ:", "", ""],
        ["• Вводите даты отпусков на листе 'СОТРУДНИКИ'", "", ""],
        ["• График обновится автоматически", "", ""],
        ["• Формат дат: ДД.ММ.ГГГГ или ДД.ММ.ГГ", "", ""],
        ["• Для добавления сотрудников копируйте строки с формулами", "", ""],
        ["• Пустые даты игнорируются", "", ""],
    ]
    
    for row_idx, row_data in enumerate(legend_data, start=3):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws_legend.cell(row=row_idx, column=col_idx, value=cell_value)
            if row_idx in [3, 9, 15]:
                cell.font = Font(bold=True)
            if row_idx >= 10 and row_idx <= 13:
                cell.font = Font(color="006400", bold=True)  # Зеленый для динамики
            if row_idx >= 16:
                cell.font = Font(color="1F497D")  # Синий для инструкций
    
    # Настраиваем ширину
    ws_legend.column_dimensions['A'].width = 35
    ws_legend.column_dimensions['B'].width = 20
    ws_legend.column_dimensions['C'].width = 50
    
    # 9. НАСТРОЙКИ ФАЙЛА
    # Закрепляем области
    ws_schedule.freeze_panes = "D4"
    ws_employees.freeze_panes = "C3"
    
    # Отключаем защиту для редактирования
    for sheet in [ws_schedule, ws_employees, ws_legend]:
        sheet.protection.sheet = False  # Отключаем защиту
    
    # 10. СОХРАНЯЕМ ФАЙЛ
    print(f"\n💾 Сохраняю файл: {filename}")
    wb.save(filename)
    
    # 11. ВЫВОД ИНФОРМАЦИИ
    print("\n" + "=" * 70)
    print("✅ ФАЙЛ УСПЕШНО СОЗДАН С ДИНАМИЧЕСКИМИ ФОРМУЛАМИ!")
    print("=" * 70)
    
    print(f"\n🎯 КЛЮЧЕВЫЕ ФУНКЦИИ НОВОЙ ВЕРСИИ:")
    print(f"   ✅ ДИНАМИЧЕСКОЕ ОБНОВЛЕНИЕ - при изменении дат отпусков")
    print(f"   ✅ Автоматический пересчет формул (F9)")
    print(f"   ✅ Условное форматирование для визуализации отпусков")
    print(f"   ✅ Не нужно перезагружать файл")
    print(f"   ✅ Сохранены все функции производственного календаря")
    
    print(f"\n📊 СТАТИСТИКА:")
    print(f"   📁 Файл: {filename}")
    print(f"   👥 Сотрудников: {len(employees_data)}")
    print(f"   📅 Дней в календаре: {len(calendar)}")
    print(f"   📏 Столбцов в графике: {current_col - 1}")
    print(f"   🔗 Формул динамической связи: {len(employees_data) * (current_col - 3)}")
    
    print(f"\n🚀 КАК РАБОТАТЬ С ФАЙЛОМ:")
    print(f"   1. Откройте файл в Excel")
    print(f"   2. На листе 'СОТРУДНИКИ' измените даты отпусков")
    print(f"   3. Нажмите F9 (или измените любую ячейку)")
    print(f"   4. На листе 'ГРАФИК ОТПУСКОВ' увидите ОБНОВЛЕННЫЕ отпуска")
    
    print(f"\n💡 ВАЖНО:")
    print(f"   • Формулы работают автоматически")
    print(f"   • Условное форматирование красит ячейки")
    print(f"   • Пустые даты на листе СОТРУДНИКИ игнорируются")
    print(f"   • Можно добавлять новых сотрудников (копируйте строки)")
    
    print(f"\n🔄 Запустите скрипт снова для создания нового файла!")
    
    return filename

def main():
    """Главная функция"""
    try:
        print("🚀 ГЕНЕРАТОР ГРАФИКОВ ОТПУСКОВ 2026")
        print("ВЕРСИЯ: ДИНАМИЧЕСКИЕ ФОРМУЛЫ")
        print("=" * 70)
        create_dynamic_vacation_schedule()
        
        # Цикл создания файлов
        while True:
            print("\n" + "-" * 70)
            another = input("\nСоздать еще один файл? (y/n): ").lower().strip()
            
            if another == 'y':
                print("\n" + "=" * 70)
                create_dynamic_vacation_schedule()
            elif another == 'n':
                print("\n👋 Завершение работы. Удачи в планировании отпусков!")
                break
            else:
                print("❌ Пожалуйста, введите 'y' или 'n'")
                
    except KeyboardInterrupt:
        print("\n\n👋 Прервано пользователем")
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        input("\nНажмите Enter для выхода...")

if __name__ == "__main__":
    # Проверяем наличие openpyxl
    try:
        from openpyxl import Workbook
    except ImportError:
        print("❌ Библиотека openpyxl не установлена!")
        print("📦 Установите командой: pip install openpyxl")
        input("\nНажмите Enter для выхода...")
        sys.exit(1)
    
    main()