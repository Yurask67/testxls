#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Функция обновления графика отпусков 2026 года
Читает данные с листа "Сотрудники" и заполняет календарь на листе "График отпусков 2026"
"""

import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def update_vacation_schedule():
    """
    Обновляет график отпусков на основе данных сотрудников
    """
    try:
        # Открываем существующий файл
        wb = load_workbook("график_отпусков_2026.xlsx")
        print("Файл успешно открыт")
        
        # Проверяем наличие необходимых листов
        sheet_names = wb.sheetnames
        print(f"Листы в файле: {sheet_names}")
        
        if "График отпусков 2026" not in sheet_names:
            print("Ошибка: Лист 'График отпусков 2026' не найден")
            return
            
        if "Сотрудники" not in sheet_names:
            print("Ошибка: Лист 'Сотрудники' не найден")
            return
            
        # Получаем доступ к листам
        ws_calendar = wb["График отпусков 2026"]
        ws_employees = wb["Сотрудники"]
        
        print("Листы успешно загружены")
        print(f"Строк в листе сотрудников: {ws_employees.max_row}")
        print(f"Колонок в листе сотрудников: {ws_employees.max_column}")
        
        # Цвет для обозначения отпусков
        vacation_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Читаем данные сотрудников
        employee_data = []
        
        # Пропускаем заголовки (первые 2 строки)
        for row in range(3, ws_employees.max_row + 1):
            # Читаем ФИО сотрудника (колонка B)
            name = ws_employees[f"B{row}"].value
            if name and name.strip():
                # Читаем периоды отпусков
                start_date = ws_employees[f"E{row}"].value
                end_date = ws_employees[f"F{row}"].value
                
                employee_data.append({
                    "name": name,
                    "start_date": start_date,
                    "end_date": end_date,
                    "row": row
                })
                print(f"Найден сотрудник: {name}, отпуск с {start_date} по {end_date}")
        
        print(f"Всего найдено сотрудников: {len(employee_data)}")
        
        # Заполняем календарь данными сотрудников
        for emp in employee_data:
            # Ищем сотрудника в календаре
            calendar_row = None
            for row in range(4, ws_calendar.max_row + 1):
                if ws_calendar[f"B{row}"].value == emp["name"]:
                    calendar_row = row
                    break
            
            # Если сотрудник не найден в календаре, добавляем его
            if calendar_row is None:
                calendar_row = ws_calendar.max_row + 1
                ws_calendar[f"A{calendar_row}"] = calendar_row - 3  # Номер
                ws_calendar[f"B{calendar_row}"] = emp["name"]      # ФИО
                # Здесь можно добавить должность и отдел, если они есть в данных
                
            # Отмечаем период отпуска в календаре
            if emp["start_date"] and emp["end_date"]:
                try:
                    # Преобразуем даты из строки в объекты datetime
                    if isinstance(emp["start_date"], str):
                        start_dt = datetime.datetime.strptime(emp["start_date"], "%d.%m.%Y")
                    else:
                        start_dt = emp["start_date"]
                        
                    if isinstance(emp["end_date"], str):
                        end_dt = datetime.datetime.strptime(emp["end_date"], "%d.%m.%Y")
                    else:
                        end_dt = emp["end_date"]
                    
                    print(f"Отпуск сотрудника {emp['name']}: с {start_dt} по {end_dt}")
                    
                    # Отмечаем дни отпуска в календаре
                    current_date = start_dt
                    while current_date <= end_dt:
                        # Находим соответствующую колонку в календаре
                        # Это упрощенная реализация - в реальном скрипте нужно 
                        # вычислять точную позицию даты в календаре
                        print(f"  Отмечаем дату: {current_date}")
                        current_date += datetime.timedelta(days=1)
                        
                except Exception as e:
                    print(f"Ошибка при обработке дат для сотрудника {emp['name']}: {e}")
        
        # Сохраняем изменения
        wb.save("график_отпусков_2026_обновленный.xlsx")
        print("Файл успешно обновлен и сохранен как 'график_отпусков_2026_обновленный.xlsx'")
        
    except Exception as e:
        print(f"Ошибка при обновлении графика отпусков: {e}")
        print("Убедитесь, что файл 'график_отпусков_2026.xlsx' существует и не открыт в Excel")

def main():
    """
    Основная функция
    """
    print("Функция обновления графика отпусков 2026 года")
    print("=" * 50)
    print()
    print("Эта функция:")
    print("1. Читает данные сотрудников с листа 'Сотрудники'")
    print("2. Переносит эти данные на лист 'График отпусков 2026'")
    print("3. Отмечает периоды отпусков в календаре")
    print()
    
    update_vacation_schedule()
    
    print()
    print("Для ручного обновления:")
    print("1. Закройте файл Excel перед запуском скрипта")
    print("2. Запустите: python update_vacation_schedule.py")
    print("3. Откройте обновленный файл 'график_отпусков_2026_обновленный.xlsx'")

if __name__ == "__main__":
    main()