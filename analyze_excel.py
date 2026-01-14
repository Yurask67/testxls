import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def analyze_excel_structure(file_path):
    """
    Анализирует структуру Excel файла
    """
    print(f"Анализ файла: {file_path}")
    
    # Загружаем книгу
    wb = load_workbook(file_path)
    print(f"Листы в файле: {wb.sheetnames}")
    
    # Анализируем каждый лист
    for sheet_name in wb.sheetnames:
        print(f"\n=== Лист: {sheet_name} ===")
        ws = wb[sheet_name]
        
        # Получаем размеры
        print(f"Максимальная строка: {ws.max_row}")
        print(f"Максимальная колонка: {ws.max_column}")
        
        # Показываем первые 10 строк и 10 колонок
        print("\nПервые 10x10 ячеек:")
        for row in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(11, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                row_data.append(f"{cell.value or ''}")
            print(f"Строка {row}: {row_data}")
        
        # Проверяем объединенные ячейки
        if ws.merged_cells.ranges:
            print(f"\nОбъединенные ячейки:")
            for merged_range in ws.merged_cells.ranges:
                print(f"  {merged_range}")

if __name__ == "__main__":
    analyze_excel_structure("график отпусков 2025.xlsx")